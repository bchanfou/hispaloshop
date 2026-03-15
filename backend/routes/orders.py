"""
Payment system, checkout, Stripe webhooks, order management,
email notifications, financial ledger, commission audit.
Extracted from server.py.
"""
from fastapi import APIRouter, HTTPException, Depends, Request, BackgroundTasks
from typing import Any, Dict, Optional, List
from datetime import datetime, timezone, timedelta
import uuid
import os
import json as _json
import logging

import stripe

from core.database import db
from core.models import (
    User, OrderCreateInput, BuyNowInput, OrderStatusUpdate,
)
from core.constants import SUPPORTED_COUNTRIES
from core.config import STRIPE_SECRET_KEY, STRIPE_WEBHOOK_SECRET
from core.auth import get_current_user, require_role
from services.auth_helpers import send_email
from services.ledger import write_ledger_event
from services.markets import get_product_target_markets, is_product_available_in_country, normalize_market_code
from config import normalize_influencer_tier
from services.shipping_service import ShippingPolicy, ShippingService

logger = logging.getLogger(__name__)
router = APIRouter()

# Set Stripe API key at module level so all calls in this module use it
stripe.api_key = STRIPE_SECRET_KEY

FRONTEND_URL = os.environ.get('FRONTEND_URL', 'http://localhost:3000')


def _round_money(amount: float) -> float:
    return round(float(amount or 0), 2)


def _first_influencer_split(commission_data: dict) -> Optional[dict]:
    for split in commission_data.get("splits", []):
        if split.get("influencer_cut", 0) > 0:
            return split
    return None


def _commission_data_to_split_details(commission_data: dict, sellers: Dict[str, Dict[str, Any]]) -> List[dict]:
    split_details = []
    for split in commission_data.get("splits", []):
        seller_meta = sellers.get(split["seller_id"], {})
        split_details.append({
            "producer_id": split["seller_id"],
            "gross_amount": _round_money(split.get("seller_gmv", 0)),
            "seller_amount": _round_money(split.get("seller_payout", 0)),
            "platform_fee": _round_money(split.get("platform_gross", 0)),
            "platform_net": _round_money(split.get("platform_net", 0)),
            "influencer_cut": _round_money(split.get("influencer_cut", 0)),
            "seller_plan": split.get("seller_plan", "FREE"),
            "commission_rate": split.get("platform_rate_snapshot", 0),
            "influencer_rate": split.get("influencer_rate_snapshot", 0),
            "influencer_tier": split.get("influencer_tier_snapshot"),
            "stripe_account_id": seller_meta.get("stripe_account_id"),
            "charges_enabled": seller_meta.get("charges_enabled", False),
        })
    return split_details


async def _get_order_commission_data(order: dict) -> dict:
    existing = order.get("commission_data")
    if existing and existing.get("splits"):
        return existing

    from services.subscriptions import calculate_order_commissions

    return await calculate_order_commissions(db, order)


async def _get_active_influencer_context(customer_id: str, customer_email: str) -> Optional[dict]:
    customer_doc = await db.users.find_one(
        {"user_id": customer_id},
        {"_id": 0, "referred_by": 1, "referral_code": 1, "referral_expires_at": 1},
    )
    if customer_doc and customer_doc.get("referred_by"):
        expiry_str = customer_doc.get("referral_expires_at")
        referral_active = True
        if expiry_str:
            try:
                referral_active = datetime.fromisoformat(expiry_str.replace("Z", "+00:00")) > datetime.now(timezone.utc)
            except Exception:
                referral_active = False

        if referral_active:
            influencer = await db.influencers.find_one(
                {"influencer_id": customer_doc["referred_by"], "status": "active"},
                {"_id": 0, "influencer_id": 1, "current_tier": 1, "commission_rate": 1, "email": 1},
            )
            if influencer and influencer.get("email") != customer_email:
                return {
                    "influencer_id": influencer["influencer_id"],
                    "discount_code": customer_doc.get("referral_code"),
                    "tier": normalize_influencer_tier(
                        influencer.get("current_tier", "hercules"),
                        influencer.get("commission_rate"),
                    ),
                }

    return None


async def _ensure_influencer_commission_record(order: dict, commission_data: dict) -> None:
    influencer_id = order.get("influencer_id")
    if not influencer_id:
        return

    influencer_amount = _round_money(commission_data.get("total_influencer_cut", 0))
    if influencer_amount <= 0:
        return

    # Fiscal gate: do not assign commission if influencer is blocked
    influencer_doc = await db.influencers.find_one(
        {"influencer_id": influencer_id}, {"_id": 0, "fiscal_status": 1}
    )
    if influencer_doc and influencer_doc.get("fiscal_status", {}).get("affiliate_blocked", False):
        logger.warning(
            f"Commission blocked for influencer {influencer_id} on order {order.get('order_id')}: "
            f"fiscal certificate not verified"
        )
        return

    existing = await db.influencer_commissions.find_one(
        {"order_id": order["order_id"], "influencer_id": influencer_id},
        {"_id": 0, "commission_id": 1},
    )
    if existing:
        return

    created_at = datetime.now(timezone.utc)
    payment_available_date = created_at + timedelta(days=15)
    first_split = _first_influencer_split(commission_data) or {}
    order_value = _round_money(sum(split.get("seller_gmv", 0) for split in commission_data.get("splits", [])))
    platform_fee = _round_money(commission_data.get("total_platform_gross", 0))

    commission_record = {
        "commission_id": f"comm_{uuid.uuid4().hex[:12]}",
        "influencer_id": influencer_id,
        "order_id": order["order_id"],
        "discount_code": order.get("influencer_discount_code"),
        "order_total": _round_money(order.get("total_amount", 0)),
        "order_value": order_value,
        "platform_fee": platform_fee,
        "commission_amount": influencer_amount,
        "commission_rate": first_split.get("influencer_rate_snapshot"),
        "influencer_tier": first_split.get("influencer_tier_snapshot"),
        "commission_status": "pending",
        "created_at": created_at.isoformat(),
        "payment_available_date": payment_available_date.isoformat(),
    }
    await db.influencer_commissions.insert_one(commission_record)

    await db.orders.update_one(
        {"order_id": order["order_id"]},
        {"$set": {
            "influencer_commission_amount": influencer_amount,
            "influencer_commission_status": "pending",
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }},
    )

    await db.influencers.update_one(
        {"influencer_id": influencer_id},
        {"$inc": {
            "total_sales_generated": order_value,
            "total_commission_earned": influencer_amount,
            "available_balance": influencer_amount,
        }},
    )

    await check_and_notify_influencer_withdrawal_available(influencer_id, db)

# ============================================================================
# PAYMENT SYSTEM — Separate Charges & Transfers Architecture
# ============================================================================
# Money flow: Customer → Platform (full amount) → Seller transfers (82%) + Influencer scheduled (tier-based, attribution 18 months)
# ============================================================================

async def _resolve_seller_stripe(seller_id: str) -> dict:
    """Resolve Stripe Connect account for a seller (producer or influencer)."""
    doc = await db.influencers.find_one({"user_id": seller_id}, {"_id": 0, "stripe_account_id": 1, "stripe_onboarding_complete": 1})
    if not doc:
        doc = await db.stores.find_one({"user_id": seller_id}, {"_id": 0, "stripe_account_id": 1, "stripe_charges_enabled": 1})
    if not doc:
        doc = await db.users.find_one({"user_id": seller_id}, {"_id": 0, "stripe_account_id": 1})
    if doc and doc.get("stripe_account_id"):
        return {
            "stripe_account_id": doc["stripe_account_id"],
            "charges_enabled": doc.get("stripe_charges_enabled", doc.get("stripe_onboarding_complete", False))
        }
    return {"stripe_account_id": None, "charges_enabled": False}


async def execute_seller_transfers(order: dict):
    """
    Execute Stripe transfers to each seller after payment is confirmed.
    Uses dynamic commission rates based on seller plan (FREE=20%, PRO=18%, ELITE=17%).
    """
    order_id = order["order_id"]
    
    if order.get("transfers_executed"):
        logger.info(f"[TRANSFERS] Already executed for {order_id}, skipping")
        return {"transfer_records": order.get("transfer_records", []), "total_platform_fee": order.get("total_platform_fee", 0)}
    
    # Use the snapshot captured at checkout when available.
    commission_data = await _get_order_commission_data(order)
    splits = commission_data.get("splits", [])
    
    currency = order.get("currency", "EUR").lower()
    transfer_records = []
    total_platform_fee = 0
    total_influencer_cut = 0
    
    for split in splits:
        seller_id = split["seller_id"]
        seller_payout_cents = int(split["seller_payout"] * 100)
        total_platform_fee += split["platform_gross"]
        total_influencer_cut += split.get("influencer_cut", 0)
        
        # Resolve Stripe account
        resolved = await _resolve_seller_stripe(seller_id)
        stripe_account_id = resolved.get("stripe_account_id")
        charges_enabled = resolved.get("charges_enabled", False)
        
        if not stripe_account_id or not charges_enabled or seller_payout_cents <= 0:
            transfer_records.append({
                "seller_id": seller_id,
                "status": "skipped",
                "reason": "no_stripe_account" if not stripe_account_id else "not_enabled",
                "amount": split["seller_payout"],
                "commission_snapshot": split,
            })
            continue
        
        try:
            transfer = stripe.Transfer.create(
                amount=seller_payout_cents,
                currency=currency,
                destination=stripe_account_id,
                transfer_group=order_id,
                metadata={
                    "order_id": order_id,
                    "seller_id": seller_id,
                    "seller_plan": split.get("seller_plan", "FREE"),
                    "platform_rate": str(split.get("platform_rate_snapshot", 0)),
                },
                idempotency_key=f"{order_id}:{seller_id}:seller_transfer"
            )
            logger.info(f"[TRANSFERS] OK: {seller_payout_cents}c → {stripe_account_id} (plan={split.get('seller_plan')}, rate={split.get('platform_rate_snapshot')})")
            transfer_records.append({
                "seller_id": seller_id,
                "transfer_id": transfer.id,
                "amount": split["seller_payout"],
                "status": "completed",
                "commission_snapshot": split,
            })
        except stripe.error.StripeError as e:
            logger.error(f"[TRANSFERS] Failed for {seller_id}: {e}")
            transfer_records.append({
                "seller_id": seller_id,
                "status": "failed",
                "error": str(e),
                "amount": split["seller_payout"],
                "commission_snapshot": split,
            })
    
    # Mark transfers as executed
    await db.orders.update_one(
        {"order_id": order_id},
        {"$set": {
            "transfers_executed": True,
            "transfer_records": transfer_records,
            "total_platform_fee": round(total_platform_fee, 2),
            "total_influencer_cut": round(total_influencer_cut, 2),
            "commission_data": commission_data,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Write ledger events for each seller transfer
    buyer_country = order.get("country", "")
    buyer_id = order.get("user_id", "")
    order_split_details = order.get("split_details", [])
    for rec in transfer_records:
        seller_split = next((s for s in order_split_details if s.get("producer_id") == rec["seller_id"]), {})
        await write_ledger_event(
            db,
            event_type="seller_transfer",
            order_id=order_id,
            currency=currency,
            product_subtotal=seller_split.get("gross_amount", 0),
            platform_fee=seller_split.get("platform_fee", 0),
            seller_net=rec.get("amount", 0),
            seller_id=rec["seller_id"],
            buyer_id=buyer_id,
            buyer_country=buyer_country,
            transfer_id=rec.get("transfer_id", ""),
            status=rec["status"],
        )
    
    return {"transfer_records": transfer_records, "total_platform_fee": total_platform_fee}


async def schedule_influencer_payout(order: dict, total_platform_fee: float, commission_data: Optional[dict] = None):
    """Schedule an influencer payout capped by the platform commission snapshot."""
    influencer_id = order.get("influencer_id")
    if not influencer_id:
        return
    
    order_id = order["order_id"]
    currency = order.get("currency", "EUR")
    
    existing = await db.scheduled_payouts.find_one({"order_id": order_id, "influencer_id": influencer_id})
    if existing:
        return

    commission_data = commission_data or await _get_order_commission_data(order)
    first_split = _first_influencer_split(commission_data)
    if not first_split:
        return

    influencer_amount = _round_money(commission_data.get("total_influencer_cut", 0))
    if influencer_amount <= 0:
        return

    inf_doc = await db.influencers.find_one(
        {"influencer_id": influencer_id},
        {"_id": 0, "current_tier": 1, "commission_rate": 1, "stripe_account_id": 1},
    )
    if not inf_doc:
        return

    inf_rate = first_split.get("influencer_rate_snapshot", inf_doc.get("commission_rate", 0.03))
    net_gmv = _round_money(sum(split.get("seller_gmv", 0) for split in commission_data.get("splits", [])))
    stripe_account_id = inf_doc.get("stripe_account_id")
    now = datetime.now(timezone.utc)
    due_date = now + timedelta(days=15)
    
    payout_record = {
        "payout_id": f"payout_{uuid.uuid4().hex[:12]}",
        "influencer_id": influencer_id,
        "influencer_stripe_account_id": stripe_account_id,
        "influencer_tier": first_split.get("influencer_tier_snapshot") or normalize_influencer_tier(
            inf_doc.get("current_tier", "hercules"),
            inf_doc.get("commission_rate"),
        ),
        "influencer_rate": inf_rate,
        "order_id": order_id,
        "amount": influencer_amount,
        "net_gmv": net_gmv,
        "platform_fee_cap": _round_money(total_platform_fee),
        "currency": currency,
        "due_date": due_date.isoformat(),
        "status": "scheduled",
        "created_at": now.isoformat()
    }
    await db.scheduled_payouts.insert_one(payout_record)
    
    # Ledger event for scheduled influencer payout
    await write_ledger_event(
        db,
        event_type="influencer_scheduled",
        order_id=order_id,
        currency=currency,
        influencer_amount=influencer_amount,
        influencer_id=influencer_id,
        buyer_id=order.get("user_id", ""),
        buyer_country=order.get("country", ""),
        platform_fee=total_platform_fee,
        status="scheduled",
    )
    
    logger.info(f"[INFLUENCER] Scheduled payout: {influencer_amount} {currency} for {influencer_id} on {due_date.date()}")


async def process_payment_confirmed(session_id: str, user_id: str = None):
    """
    Central post-payment processing. Called by webhook or checkout-status polling.
    Handles: order status, stock decrement, transfers, notifications, influencer scheduling.
    """
    order = await db.orders.find_one({"payment_session_id": session_id}, {"_id": 0})
    if not order:
        # Try pending_orders (buy-now flow)
        pending = await db.pending_orders.find_one({"session_id": session_id}, {"_id": 0})
        if pending:
            # Promote pending order to real order
            pending["payment_session_id"] = session_id
            pending["status"] = "paid"
            pending["updated_at"] = datetime.now(timezone.utc).isoformat()
            pending.pop("session_id", None)
            await db.orders.insert_one(pending)
            await db.pending_orders.delete_one({"order_id": pending["order_id"]})
            order = await db.orders.find_one({"payment_session_id": session_id}, {"_id": 0})
        if not order:
            logger.error(f"[PAYMENT] No order found for session {session_id}")
            return
    
    if order.get("status") in ("paid", "confirmed", "preparing", "shipped", "delivered"):
        if order.get("transfers_executed"):
            logger.info(f"[PAYMENT] Order {order['order_id']} already fully processed")
            return
    
    order_id = order["order_id"]
    
    # 1. Mark order as paid
    await db.payment_transactions.update_one(
        {"session_id": session_id},
        {"$set": {"status": "paid", "payment_status": "paid", "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    await db.orders.update_one(
        {"order_id": order_id},
        {"$set": {"status": "paid", "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    # Refresh order after update
    order = await db.orders.find_one({"order_id": order_id}, {"_id": 0})
    
    # 2. Atomic stock decrement
    for item in order.get("line_items", []):
        product = await db.products.find_one(
            {"product_id": item["product_id"]},
            {"_id": 0, "track_stock": 1, "variants": 1}
        )
        if product and product.get("track_stock", True):
            variant_id = item.get("variant_id")
            pack_id = item.get("pack_id")
            if variant_id and pack_id and product.get("variants"):
                await db.products.update_one(
                    {"product_id": item["product_id"], "variants.variant_id": variant_id, "variants.packs.pack_id": pack_id},
                    {"$inc": {"variants.$[v].packs.$[p].stock": -item["quantity"]}},
                    array_filters=[{"v.variant_id": variant_id}, {"p.pack_id": pack_id}]
                )
            else:
                await db.products.update_one(
                    {"product_id": item["product_id"], "stock": {"$gte": item["quantity"]}},
                    {"$inc": {"stock": -item["quantity"]}}
                )
    
    # 3. Increment discount usage
    if order.get("discount_info"):
        code = order["discount_info"].get("code")
        if code:
            await db.discount_codes.update_one({"code": code}, {"$inc": {"usage_count": 1}})
    
    # 4. Clear cart
    if user_id:
        await db.cart_items.delete_many({"user_id": user_id})
        await db.cart_discounts.delete_one({"user_id": user_id})

    # 4a. Release soft-holds (stock reservations)
    await db.stock_holds.delete_many({"session_id": session_id})
    if user_id:
        await db.stock_holds.delete_many({"user_id": user_id})

    # 4b. Write order_paid ledger event
    await write_ledger_event(
        db,
        event_type="order_paid",
        order_id=order_id,
        currency=order.get("currency", "EUR"),
        product_subtotal=order.get("total_amount", 0),
        buyer_id=order.get("user_id", ""),
        buyer_country=order.get("country", ""),
        status="completed",
    )
    
    commission_data = await _get_order_commission_data(order)

    # 5. Execute seller transfers (Separate Charges & Transfers)
    transfer_result = await execute_seller_transfers(order)
    total_platform_fee = transfer_result["total_platform_fee"] if transfer_result else 0

    # 6. Persist influencer commission only after payment confirmation.
    await _ensure_influencer_commission_record(order, commission_data)

    # 7. Schedule influencer payout (15 days deferred)
    await schedule_influencer_payout(order, total_platform_fee, commission_data)

    # 8. Notify producers + consumer via unified notification service
    from routes.notifications import notify_order_event
    try:
        await notify_order_event(order_id, "order_confirmed")
    except Exception as e:
        logger.error(f"[NOTIF] Failed notify_order_event for {order_id}: {e}")

    producers = set(item["producer_id"] for item in order.get("line_items", []))
    for producer_id in producers:
        producer_items = [item for item in order["line_items"] if item["producer_id"] == producer_id]
        try:
            await send_new_order_email_to_producer(producer_id, order, producer_items)
        except Exception as e:
            logger.error(f"[EMAIL] Failed for producer {producer_id}: {e}")
    
    # 9. Send confirmation email to customer
    try:
        await send_order_status_email(order, "confirmed")
    except Exception as e:
        logger.error(f"[EMAIL] Failed customer confirmation for order {order_id}: {e}")

    # 10. Create commission_transaction record for audit trail
    from services.ledger import EXCHANGE_RATES_TO_USD
    currency = order.get("currency", "EUR")
    usd_rate = EXCHANGE_RATES_TO_USD.get(currency.upper(), 1.0)
    
    commission_splits = transfer_result.get("transfer_records", []) if transfer_result else []
    for rec in commission_splits:
        snap = rec.get("commission_snapshot", {})
        await db.commission_transactions.insert_one({
            "transaction_id": f"ctx_{uuid.uuid4().hex[:12]}",
            "order_id": order_id,
            "event_type": "order_paid",
            "seller_id": rec.get("seller_id", ""),
            "influencer_id": snap.get("influencer_id"),
            "net_gmv_usd": round(snap.get("seller_gmv", 0) * usd_rate, 2),
            "platform_rate_applied": snap.get("platform_rate_snapshot"),
            "platform_earnings_usd": round(snap.get("platform_net", 0) * usd_rate, 2),
            "influencer_rate_applied": snap.get("influencer_rate_snapshot"),
            "influencer_earnings_usd": round(snap.get("influencer_cut", 0) * usd_rate, 2) if snap.get("influencer_cut") else None,
            "seller_payout_usd": round(snap.get("seller_payout", 0) * usd_rate, 2),
            "currency_original": currency,
            "exchange_rate_to_usd": usd_rate,
            "status": rec.get("status", "pending"),
            "seller_payout_date": datetime.now(timezone.utc).isoformat() if rec.get("status") == "completed" else None,
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
    
    logger.info(f"[PAYMENT] Order {order_id} fully processed: transfers={'done' if transfer_result else 'skipped'}")


async def process_influencer_scheduled_payouts():
    """
    Daily job: execute scheduled influencer payouts that are due.
    Call this from a background task or cron.
    """
    now = datetime.now(timezone.utc).isoformat()
    due_payouts = await db.scheduled_payouts.find(
        {"status": "scheduled", "due_date": {"$lte": now}}
    ).to_list(100)
    
    for payout in due_payouts:
        payout_id = payout["payout_id"]
        stripe_account_id = payout.get("influencer_stripe_account_id")
        amount_cents = int(payout["amount"] * 100)
        currency = payout.get("currency", "eur").lower()
        order_id = payout["order_id"]
        
        # Verify order is still paid (not refunded)
        order = await db.orders.find_one({"order_id": order_id}, {"_id": 0, "status": 1})
        if not order or order.get("status") in ("cancelled", "refunded"):
            await db.scheduled_payouts.update_one(
                {"payout_id": payout_id},
                {"$set": {"status": "cancelled", "cancel_reason": "order_cancelled", "updated_at": now}}
            )
            continue
        
        if not stripe_account_id or amount_cents <= 0:
            await db.scheduled_payouts.update_one(
                {"payout_id": payout_id},
                {"$set": {"status": "skipped", "skip_reason": "no_stripe_account", "updated_at": now}}
            )
            continue
        
        try:
            transfer = stripe.Transfer.create(
                amount=amount_cents,
                currency=currency,
                destination=stripe_account_id,
                transfer_group=order_id,
                metadata={"order_id": order_id, "influencer_id": payout["influencer_id"], "type": "influencer_payout"},
                idempotency_key=f"{order_id}:{payout['influencer_id']}:influencer_payout"
            )
            await db.scheduled_payouts.update_one(
                {"payout_id": payout_id},
                {"$set": {"status": "paid", "transfer_id": transfer.id, "paid_at": now, "updated_at": now}}
            )
            # Ledger event
            await write_ledger_event(
                db,
                event_type="influencer_paid",
                order_id=order_id,
                currency=currency,
                influencer_amount=payout["amount"],
                influencer_id=payout["influencer_id"],
                transfer_id=transfer.id,
                status="completed",
            )
            logger.info(f"[INFLUENCER-PAYOUT] Paid {payout['amount']} {currency} to {payout['influencer_id']} (transfer {transfer.id})")
        except stripe.error.StripeError as e:
            logger.error(f"[INFLUENCER-PAYOUT] Failed for {payout_id}: {e}")
            await db.scheduled_payouts.update_one(
                {"payout_id": payout_id},
                {"$set": {"status": "failed", "error": str(e), "updated_at": now}}
            )


# ── Payment Routes ───────────────────────────────────────────

@router.get("/payments/stripe-status")
async def stripe_status():
    """Health check: returns whether Stripe is properly configured."""
    key = STRIPE_SECRET_KEY or ""
    is_live = key.startswith("sk_live_") and len(key) > 20 and "PENDIENTE" not in key
    is_test = key.startswith("sk_test_") and len(key) > 20
    webhook = STRIPE_WEBHOOK_SECRET or ""
    webhook_ok = webhook.startswith("whsec_") and len(webhook) > 32 and "PENDIENTE" not in webhook
    return {
        "stripe_configured": is_live or is_test,
        "mode": "live" if is_live else ("test" if is_test else "not_configured"),
        "webhook_configured": webhook_ok,
    }


@router.post("/payments/create-checkout")
async def create_checkout(request: Request, input: OrderCreateInput, user: User = Depends(get_current_user)):
    # Guard: Stripe must be configured with a real key
    _sk = STRIPE_SECRET_KEY or ""
    if not (_sk.startswith(("sk_live_", "sk_test_")) and len(_sk) > 20 and "PENDIENTE" not in _sk):
        raise HTTPException(status_code=503, detail="Payment processing not configured. Contact the administrator.")

    # Check email verification before checkout
    user_doc = await db.users.find_one({"user_id": user.user_id}, {"_id": 0, "email_verified": 1, "locale": 1})
    if not user_doc.get("email_verified", False):
        raise HTTPException(
            status_code=403, 
            detail="Please verify your email before placing an order. Check your email or request a new verification link."
        )
    
    # Get user's selected country for pricing and currency
    user_country = normalize_market_code(user_doc.get("locale", {}).get("country")) or "ES"
    base_currency = SUPPORTED_COUNTRIES.get(user_country, {}).get("currency", "EUR")
    
    cart_items = await db.cart_items.find({"user_id": user.user_id}, {"_id": 0}).to_list(100)
    if not cart_items:
        raise HTTPException(status_code=400, detail="Cart is empty")
    
    # === COUNTRY AVAILABILITY & STOCK VALIDATION AT CHECKOUT ===
    stock_issues = []
    for item in cart_items:
        product = await db.products.find_one(
            {"product_id": item["product_id"]}, 
            {"_id": 0, "stock": 1, "track_stock": 1, "name": 1, "variants": 1, 
             "available_countries": 1, "country_prices": 1}
        )
        if not product:
            stock_issues.append(f"Product '{item['product_name']}' no longer available")
            continue
        
        # Validate country availability
        if not is_product_available_in_country(product, user_country):
            stock_issues.append(f"Product '{item['product_name']}' is not available in {user_country}")
            continue
        
        # Auto-correct pricing if changed (don't reject, just update)
        current_price = product.get("price", 0)
        inv = product.get("inventory_by_country", [])
        market = next((m for m in inv if m.get("country_code") == user_country and m.get("active")), None)
        if market:
            current_price = market.get("price", current_price)
        elif product.get("country_prices", {}).get(user_country):
            current_price = product["country_prices"][user_country]
        
        if abs(item["price"] - current_price) > 0.01:
            # Auto-update cart item price
            await db.cart_items.update_one(
                {"user_id": user.user_id, "product_id": item["product_id"]},
                {"$set": {"price": current_price}}
            )
            item["price"] = current_price
        
        track_stock = product.get("track_stock", True)
        variant_id = item.get("variant_id")
        pack_id = item.get("pack_id")
        
        # Determine stock based on variant/pack or product level
        if variant_id and pack_id and product.get("variants"):
            # Find pack stock
            current_stock = 0
            for variant in product["variants"]:
                if variant["variant_id"] == variant_id:
                    for pack in variant.get("packs", []):
                        if pack["pack_id"] == pack_id:
                            current_stock = pack.get("stock", 0)
                            break
                    break
            item_label = f"{item['product_name']} ({item.get('variant_name', '')} - {item.get('pack_label', '')})"
        else:
            current_stock = product.get("stock", 0)
            item_label = item['product_name']
        
        if track_stock:
            if current_stock <= 0:
                stock_issues.append(f"'{item_label}' is out of stock")
            elif item["quantity"] > current_stock:
                stock_issues.append(f"Only {current_stock} units of '{item_label}' available")
    
    if stock_issues:
        raise HTTPException(status_code=400, detail={"message": "Checkout validation failed", "issues": stock_issues})
    
    # Calculate base total
    subtotal = sum(item["price"] * item["quantity"] for item in cart_items)
    
    # === APPLY DISCOUNT IF ANY ===
    discount_amount = 0
    applied_discount = await db.cart_discounts.find_one({"user_id": user.user_id}, {"_id": 0})
    discount_info = None
    influencer_commission_data = None  # Phase 4: Track influencer commission
    
    if applied_discount:
        # Re-validate discount code at checkout time
        discount_code = await db.discount_codes.find_one({"code_id": applied_discount["code_id"]}, {"_id": 0})
        
        if discount_code and discount_code.get("active", True):
            now = datetime.now(timezone.utc).isoformat()
            valid = True
            
            if discount_code.get("start_date") and now < discount_code["start_date"]:
                valid = False
            if discount_code.get("end_date") and now > discount_code["end_date"]:
                valid = False
            if discount_code.get("usage_limit") is not None and discount_code.get("usage_count", 0) >= discount_code["usage_limit"]:
                valid = False
            if discount_code.get("min_cart_amount") and subtotal < discount_code["min_cart_amount"]:
                valid = False
            # Influencer coupons are first-order only for the customer using the code.
            if valid and discount_code.get("influencer_id"):
                prior_orders = await db.orders.count_documents({
                    "user_id": user.user_id,
                    "status": {"$in": ["paid", "confirmed", "preparing", "shipped", "delivered"]}
                })
                if prior_orders > 0:
                    valid = False
            
            if valid:
                # Calculate discount
                if discount_code["type"] == "percentage":
                    applicable_products = discount_code.get("applicable_products", [])
                    if applicable_products:
                        applicable_total = sum(
                            item["price"] * item["quantity"] 
                            for item in cart_items 
                            if item["product_id"] in applicable_products
                        )
                        discount_amount = applicable_total * (discount_code["value"] / 100)
                    else:
                        discount_amount = subtotal * (discount_code["value"] / 100)
                elif discount_code["type"] == "fixed":
                    discount_amount = min(discount_code["value"], subtotal)
                # free_shipping would be handled separately in shipping calculation
                
                discount_info = {
                    "code": discount_code["code"],
                    "type": discount_code["type"],
                    "value": discount_code["value"],
                    "discount_amount": round(discount_amount, 2)
                }
                
                # Attribution lock + tier-based commission
                if discount_code.get("influencer_id"):
                    from services.referrals import check_influencer_attribution, create_attribution
                    from services.subscriptions import get_influencer_commission_rate
                    attr_check = await check_influencer_attribution(db, user.user_id, discount_code["code"])
                    
                    if not attr_check["allowed"]:
                        discount_info = None
                        discount_amount = 0
                    else:
                        influencer = await db.influencers.find_one(
                            {"influencer_id": attr_check["influencer_id"], "status": "active"}, {"_id": 0}
                        )
                        if influencer and influencer.get("email") != user.email:
                            influencer_tier = normalize_influencer_tier(
                                influencer.get("current_tier", "hercules"),
                                influencer.get("commission_rate"),
                            )
                            inf_rate = influencer.get("commission_rate", get_influencer_commission_rate(influencer_tier))
                            net_order_value = subtotal - discount_amount
                            commission_amount = round(net_order_value * inf_rate, 2)
                            influencer_commission_data = {
                                "influencer_id": influencer["influencer_id"],
                                "discount_code": discount_code["code"],
                                "commission_amount": commission_amount,
                                "commission_rate": inf_rate,
                                "tier": influencer_tier,
                                "order_value": round(net_order_value, 2),
                            }
                            if not attr_check.get("existing"):
                                await create_attribution(db, user.user_id, influencer["influencer_id"], discount_code["code"])

    discounted_subtotal = max(0, subtotal - discount_amount)

    # Active referral attribution:
    # once a customer is linked through referred_by, future orders reuse that influencer
    # while the referral window is still valid.
    if not influencer_commission_data:
        influencer_context = await _get_active_influencer_context(user.user_id, user.email)
        if influencer_context:
            from services.subscriptions import get_influencer_commission_rate

            influencer_tier = influencer_context["tier"]
            inf_rate = get_influencer_commission_rate(influencer_tier)
            commission_amount = round(discounted_subtotal * inf_rate, 2)
            influencer_commission_data = {
                "influencer_id": influencer_context["influencer_id"],
                "discount_code": influencer_context.get("discount_code"),
                "commission_amount": commission_amount,
                "commission_rate": inf_rate,
                "tier": influencer_tier,
                "order_value": round(discounted_subtotal, 2),
            }

    # Shipping policy calculation by producer
    producer_groups: dict[str, dict] = {}
    for item in cart_items:
        pid = item.get("producer_id")
        if not pid:
            continue
        producer_groups.setdefault(pid, {"subtotal_cents": 0, "item_count": 0})
        producer_groups[pid]["subtotal_cents"] += int(round(item.get("price", 0) * item.get("quantity", 0) * 100))
        producer_groups[pid]["item_count"] += int(item.get("quantity", 0) or 0)

    shipping_cents = 0
    for producer_id, group in producer_groups.items():
        producer_doc = await db.users.find_one(
            {"user_id": producer_id},
            {
                "_id": 0,
                "shipping_policy_enabled": 1,
                "shipping_base_cost_cents": 1,
                "shipping_free_threshold_cents": 1,
                "shipping_per_item_cents": 1,
            },
        )
        if not producer_doc:
            continue
        policy = ShippingPolicy(
            enabled=bool(producer_doc.get("shipping_policy_enabled", False)),
            base_cost_cents=int(producer_doc.get("shipping_base_cost_cents", 0) or 0),
            per_item_cents=int(producer_doc.get("shipping_per_item_cents", 0) or 0),
            free_threshold_cents=producer_doc.get("shipping_free_threshold_cents"),
        )
        shipping_cents += ShippingService.calculate_shipping_cents(
            policy=policy,
            item_count=group["item_count"],
            subtotal_cents=group["subtotal_cents"],
        )

    if discount_info and discount_info.get("type") == "free_shipping":
        shipping_cents = 0

    tax_rate_bp = ShippingService.get_tax_rate_bp(user_country)
    totals_cents = ShippingService.calculate_order_totals(
        subtotal_cents=int(round(discounted_subtotal * 100)),
        shipping_cents=shipping_cents,
        tax_rate_bp=tax_rate_bp,
    )

    shipping_amount = round(totals_cents["shipping_cents"] / 100, 2)
    tax_amount = round(totals_cents["tax_cents"] / 100, 2)
    total_amount = round(totals_cents["total_cents"] / 100, 2)
    order_id = f"order_{uuid.uuid4().hex[:12]}"
    
    line_items = []
    producers_in_order = {}  # {producer_id: {amount, stripe_account_id}}
    for item in cart_items:
        line_item = {
            "product_id": item["product_id"],
            "product_name": item["product_name"],
            "producer_id": item["producer_id"],
            "quantity": item["quantity"],
            "price": item["price"],
            "subtotal": item["price"] * item["quantity"]
        }
        if item.get("variant_id"):
            line_item["variant_id"] = item["variant_id"]
            line_item["variant_name"] = item.get("variant_name")
        if item.get("pack_id"):
            line_item["pack_id"] = item["pack_id"]
            line_item["pack_label"] = item.get("pack_label")
            line_item["pack_units"] = item.get("pack_units")
        line_items.append(line_item)
        
        # Aggregate amounts per producer for split calculation
        pid = item["producer_id"]
        if pid not in producers_in_order:
            producers_in_order[pid] = {"amount": 0}
        producers_in_order[pid]["amount"] += item["price"] * item["quantity"]
    
    # Resolve Stripe Connect accounts for producers
    stripe_issues = []
    for pid in producers_in_order:
        resolved = await _resolve_seller_stripe(pid)
        producers_in_order[pid]["stripe_account_id"] = resolved["stripe_account_id"]
        producers_in_order[pid]["charges_enabled"] = resolved["charges_enabled"]
        if not resolved["stripe_account_id"]:
            stripe_issues.append(f"Seller {pid} has no Stripe account connected")
        elif not resolved["charges_enabled"]:
            stripe_issues.append(f"Seller {pid} has not completed Stripe onboarding")
    
    # Pre-checkout validation: all sellers must have active Stripe accounts
    if stripe_issues:
        logger.warning(f"[CHECKOUT] Stripe validation issues: {stripe_issues}")
        # Allow checkout anyway — platform collects, manual payout later
        # raise HTTPException(status_code=400, detail={"message": "Some sellers cannot receive payments", "issues": stripe_issues})
    
    order_commission_seed = {
        "order_id": order_id,
        "line_items": line_items,
        "total_amount": round(total_amount, 2),
        "influencer_id": influencer_commission_data["influencer_id"] if influencer_commission_data else None,
    }
    commission_data = await _get_order_commission_data(order_commission_seed)
    split_details = _commission_data_to_split_details(commission_data, producers_in_order)

    first_influencer_split = _first_influencer_split(commission_data)
    total_influencer_cut = _round_money(commission_data.get("total_influencer_cut", 0))
    if influencer_commission_data and first_influencer_split:
        influencer_commission_data.update({
            "commission_amount": total_influencer_cut,
            "commission_rate": first_influencer_split.get("influencer_rate_snapshot", influencer_commission_data.get("commission_rate")),
            "tier": first_influencer_split.get("influencer_tier_snapshot", influencer_commission_data.get("tier")),
            "order_value": _round_money(sum(split.get("seller_gmv", 0) for split in commission_data.get("splits", []))),
            "platform_fee": _round_money(commission_data.get("total_platform_gross", 0)),
        })
    
    host_url = str(request.base_url).rstrip('/')
    origin = request.headers.get('origin', host_url)
    success_url = f"{origin}/checkout/success?session_id={{CHECKOUT_SESSION_ID}}"
    cancel_url = f"{origin}/cart"
    
    # Serialize seller_breakdown for Stripe metadata (max 500 chars per value)
    seller_breakdown = {}
    for split in split_details:
        seller_breakdown[split["producer_id"]] = int(split["gross_amount"] * 100)
    
    # Build Stripe Checkout Session — Separate Charges & Transfers
    # Money enters platform account fully. Transfers executed after webhook confirmation.
    stripe_params = {
        "mode": "payment",
        "success_url": success_url,
        "cancel_url": cancel_url,
        "customer_email": user.email,
        "payment_intent_data": {
            "transfer_group": order_id,
        },
        "metadata": {
            "order_id": order_id,
            "user_id": user.user_id,
            "country": user_country,
            "seller_breakdown": _json.dumps(seller_breakdown),
            "influencer_id": influencer_commission_data["influencer_id"] if influencer_commission_data else "",
            "shipping_cents": str(totals_cents["shipping_cents"]),
            "tax_cents": str(totals_cents["tax_cents"]),
            "tax_rate_bp": str(tax_rate_bp),
        },
        "line_items": [{
            "price_data": {
                "currency": base_currency.lower(),
                "unit_amount": int(total_amount * 100),
                "product_data": {"name": f"Pedido Hispaloshop #{order_id[-8:]}"},
            },
            "quantity": 1,
        }],
    }
    
    try:
        session = stripe.checkout.Session.create(**stripe_params)
    except stripe.error.StripeError as e:
        logger.error(f"[CHECKOUT] Stripe error: {e}")
        raise HTTPException(status_code=500, detail="Payment processing error. Please try again.")
    
    transaction_id = f"txn_{uuid.uuid4().hex[:12]}"
    transaction = {
        "transaction_id": transaction_id,
        "session_id": session.id,
        "order_id": order_id,
        "user_id": user.user_id,
        "amount": total_amount,
        "currency": base_currency,
        "country": user_country,
        "status": "initiated",
        "payment_status": "pending",
        "split_details": split_details,
        "commission_data": commission_data,
        "metadata": {"order_id": order_id, "country": user_country},
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    await db.payment_transactions.insert_one(transaction)
    
    order = {
        "order_id": order_id,
        "user_id": user.user_id,
        "user_email": user.email,
        "user_name": user.name,
        "country": user_country,
        "currency": base_currency,
        "subtotal": subtotal,
        "subtotal_after_discount": round(discounted_subtotal, 2),
        "discount_info": discount_info,
        "discount_amount": round(discount_amount, 2),
        "shipping_amount": shipping_amount,
        "tax_amount": tax_amount,
        "tax_rate_bp": tax_rate_bp,
        "total_amount": round(total_amount, 2),
        "status": "pending",
        "line_items": line_items,
        "split_details": split_details,
        "commission_data": commission_data,
        "financial_snapshot": commission_data,
        "shipping_address": input.shipping_address,
        "payment_session_id": session.id,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "influencer_id": influencer_commission_data["influencer_id"] if influencer_commission_data else None,
        "influencer_discount_code": influencer_commission_data["discount_code"] if influencer_commission_data else None,
        "influencer_commission_amount": influencer_commission_data["commission_amount"] if influencer_commission_data else None,
        "influencer_commission_status": "pending" if influencer_commission_data else None
    }
    await db.orders.insert_one(order)
    return {"url": session.url, "session_id": session.id}

@router.post("/checkout/buy-now")
async def buy_now_checkout(input: BuyNowInput, request: Request, user: User = Depends(get_current_user)):
    """
    Create direct checkout session for Buy Now (skip cart)
    """
    _sk = STRIPE_SECRET_KEY or ""
    if not (_sk.startswith(("sk_live_", "sk_test_")) and len(_sk) > 20 and "PENDIENTE" not in _sk):
        raise HTTPException(status_code=503, detail="Payment processing not configured. Contact the administrator.")

    # Get user's country for pricing
    user_country = normalize_market_code(user.country) or 'ES'
    
    # Fetch product with country-specific pricing
    product = await db.products.find_one({"product_id": input.product_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    # Check country availability
    product["target_markets"] = get_product_target_markets(product)
    if not is_product_available_in_country(product, user_country):
        raise HTTPException(status_code=400, detail=f"Product not available in {user_country}")
    
    # Get country-specific price and currency
    country_prices = product.get("country_prices", {})
    country_currency = product.get("country_currency", {})
    
    base_price = country_prices.get(user_country, product.get("price", 0))
    base_currency = country_currency.get(user_country, "EUR")
    
    # Handle variant selection
    variant_price_modifier = 0
    variant_name = None
    if input.variant_id:
        variants = product.get("variants", [])
        variant = next((v for v in variants if v["variant_id"] == input.variant_id), None)
        if not variant:
            raise HTTPException(status_code=400, detail="Variant not found")
        variant_name = variant.get("name")
        variant_price_modifier = variant.get("price_modifier", 0)
    
    # Handle pack selection (packs live inside variants when variants exist)
    pack_price = None
    pack_label = None
    pack_units = None
    if input.pack_id:
        if input.variant_id and variant:
            packs = variant.get("packs", [])
        else:
            packs = product.get("packs", [])
        pack = next((p for p in packs if p["pack_id"] == input.pack_id), None)
        if not pack:
            raise HTTPException(status_code=400, detail="Pack not found")
        pack_price = pack.get("price")
        pack_label = pack.get("label")
        pack_units = pack.get("units")
    
    # Calculate final price
    if pack_price is not None:
        unit_price = pack_price
    else:
        unit_price = base_price + variant_price_modifier
    
    # Stock validation (check pack-level stock when applicable)
    track_stock = product.get("track_stock", True)
    if track_stock:
        if input.variant_id and input.pack_id and pack:
            current_stock = pack.get("stock", 0)
        elif input.variant_id and variant:
            current_stock = variant.get("stock", product.get("stock", 0))
        else:
            current_stock = product.get("stock", 0)
        if current_stock < input.quantity:
            raise HTTPException(
                status_code=400, 
                detail=f"Insufficient stock. Only {current_stock} units available"
            )
    
    # Calculate total
    total_amount = unit_price * input.quantity
    
    # Generate order ID
    order_id = f"order_{uuid.uuid4().hex[:12]}"
    
    # Create line item
    line_item = {
        "product_id": input.product_id,
        "product_name": product["name"],
        "producer_id": product.get("producer_id", ""),
        "quantity": input.quantity,
        "price": unit_price,
        "subtotal": total_amount
    }
    
    # Add variant/pack info
    if input.variant_id:
        line_item["variant_id"] = input.variant_id
        line_item["variant_name"] = variant_name
    if input.pack_id:
        line_item["pack_id"] = input.pack_id
        line_item["pack_label"] = pack_label
        line_item["pack_units"] = pack_units
    
    # Resolve Stripe Connect account for the producer
    producer_id = product.get("producer_id", "")
    producer_stripe = None
    if producer_id:
        producer_stripe = await db.influencers.find_one({"user_id": producer_id}, {"_id": 0, "stripe_account_id": 1, "stripe_onboarding_complete": 1})
        if not producer_stripe:
            producer_stripe = await db.stores.find_one({"user_id": producer_id}, {"_id": 0, "stripe_account_id": 1, "stripe_charges_enabled": 1})

    influencer_context = await _get_active_influencer_context(user.user_id, user.email)
    buy_now_commission_seed = {
        "order_id": order_id,
        "line_items": [line_item],
        "total_amount": round(total_amount, 2),
        "influencer_id": influencer_context["influencer_id"] if influencer_context else None,
    }
    commission_data = await _get_order_commission_data(buy_now_commission_seed)
    split_details = _commission_data_to_split_details(commission_data, {
        producer_id: {
            "stripe_account_id": producer_stripe.get("stripe_account_id") if producer_stripe else None,
            "charges_enabled": producer_stripe.get("stripe_charges_enabled", producer_stripe.get("stripe_onboarding_complete", False)) if producer_stripe else False,
        }
    })
    
    # Setup Stripe checkout — Separate Charges & Transfers
    host_url = str(request.base_url).rstrip('/')
    origin = request.headers.get('origin', host_url)
    success_url = f"{origin}/checkout/success?session_id={{CHECKOUT_SESSION_ID}}"
    cancel_url = f"{origin}/products/{input.product_id}"
    
    seller_breakdown = {producer_id: int(total_amount * 100)}
    
    stripe_params = {
        "mode": "payment",
        "success_url": success_url,
        "cancel_url": cancel_url,
        "customer_email": user.email,
        "payment_intent_data": {
            "transfer_group": order_id,
        },
        "metadata": {
            "order_id": order_id,
            "user_id": user.user_id,
            "country": user_country,
            "buy_now": "true",
            "seller_breakdown": _json.dumps(seller_breakdown),
            "influencer_id": influencer_context["influencer_id"] if influencer_context else "",
        },
        "line_items": [{
            "price_data": {
                "currency": base_currency.lower(),
                "unit_amount": int(total_amount * 100),
                "product_data": {"name": f"Compra directa #{order_id[-8:]}"},
            },
            "quantity": 1,
        }],
    }
    
    try:
        session = stripe.checkout.Session.create(**stripe_params)
    except stripe.error.StripeError as e:
        logger.error(f"[BUY_NOW] Stripe error: {e}")
        raise HTTPException(status_code=500, detail="Payment processing error. Please try again.")
    
    # Store transaction
    transaction_id = f"txn_{uuid.uuid4().hex[:12]}"
    transaction = {
        "transaction_id": transaction_id,
        "session_id": session.id,
        "order_id": order_id,
        "user_id": user.user_id,
        "amount": total_amount,
        "currency": base_currency,
        "country": user_country,
        "status": "initiated",
        "payment_status": "pending",
        "split_details": split_details,
        "commission_data": commission_data,
        "metadata": {"order_id": order_id, "country": user_country, "buy_now": True},
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    await db.payment_transactions.insert_one(transaction)
    
    # Store pending order
    pending_order = {
        "order_id": order_id,
        "user_id": user.user_id,
        "session_id": session.id,
        "line_items": [line_item],
        "subtotal": total_amount,
        "total_amount": total_amount,
        "currency": base_currency,
        "country": user_country,
        "discount_info": None,
        "split_details": split_details,
        "commission_data": commission_data,
        "financial_snapshot": commission_data,
        "influencer_id": influencer_context["influencer_id"] if influencer_context else None,
        "influencer_discount_code": influencer_context["discount_code"] if influencer_context else None,
        "influencer_commission_amount": _round_money(commission_data.get("total_influencer_cut", 0)) if influencer_context else None,
        "influencer_commission_status": "pending" if influencer_context and commission_data.get("total_influencer_cut", 0) > 0 else None,
        "status": "pending_payment",
        "buy_now": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.pending_orders.insert_one(pending_order)
    
    logger.info(f"[BUY_NOW] Created checkout session for user {user.user_id}, order {order_id}")
    
    return {
        "checkout_url": session.url,
        "session_id": session.id,
        "order_id": order_id
    }

@router.get("/payments/checkout-status/{session_id}")
async def checkout_status(session_id: str, user: User = Depends(get_current_user)):
    """Poll payment status. Falls back from webhook if webhook hasn't fired yet."""
    try:
        session = stripe.checkout.Session.retrieve(session_id)
    except stripe.error.StripeError as e:
        logger.error(f"[CHECKOUT-STATUS] Stripe error: {e}")
        raise HTTPException(status_code=400, detail="Could not retrieve payment status")
    
    transaction = await db.payment_transactions.find_one({"session_id": session_id}, {"_id": 0})
    if not transaction:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    if transaction["status"] == "paid":
        return {"message": "Payment already processed", "status": "paid"}
    
    payment_status = session.payment_status
    
    if payment_status == "paid":
        # Trigger full post-payment processing (idempotent)
        await process_payment_confirmed(session_id, user_id=user.user_id)
    
    return {"status": session.status, "payment_status": payment_status, "amount_total": session.amount_total, "currency": session.currency}


@router.get("/checkout/{checkout_id}")
async def legacy_checkout_status(checkout_id: str, user: User = Depends(get_current_user)):
    """Legacy alias for checkout status polling used by older frontend hooks."""
    return await checkout_status(checkout_id, user)


@router.post("/checkout/{checkout_id}/confirm")
async def legacy_checkout_confirm(checkout_id: str, user: User = Depends(get_current_user)):
    """Legacy alias to force checkout processing for older frontend flows."""
    await process_payment_confirmed(checkout_id, user_id=user.user_id)
    return {"success": True, "status": "processed"}

@router.post("/webhook/stripe")
async def stripe_webhook(request: Request):
    """
    Stripe webhook — primary payment confirmation handler.
    Validates signature, processes payment, executes seller transfers.
    """
    body = await request.body()
    signature = request.headers.get("Stripe-Signature")
    
    # A real webhook secret starts with "whsec_" and is ≥32 chars after that prefix.
    _real_webhook_secret = (
        STRIPE_WEBHOOK_SECRET
        if (STRIPE_WEBHOOK_SECRET and STRIPE_WEBHOOK_SECRET.startswith("whsec_") and len(STRIPE_WEBHOOK_SECRET) > 32)
        else None
    )

    try:
        # Verify webhook signature if a real secret is configured
        if _real_webhook_secret:
            event = stripe.Webhook.construct_event(body, signature, _real_webhook_secret)
        else:
            # No real webhook secret: parse payload directly (test/dev mode — signature not verified)
            logger.warning("[WEBHOOK] Signature verification skipped — STRIPE_WEBHOOK_SECRET not configured")
            payload = body.decode("utf-8")
            event = _json.loads(payload)
        
        event_type = event.get("type", event.get("type", "unknown"))
        logger.info(f"[WEBHOOK] Event: {event_type}")
        
        if event_type == "checkout.session.completed":
            session_obj = event.get("data", {}).get("object", {})
            session_id = session_obj.get("id")
            payment_status = session_obj.get("payment_status")
            metadata = session_obj.get("metadata", {})
            user_id = metadata.get("user_id")
            
            if payment_status == "paid" and session_id:
                logger.info(f"[WEBHOOK] Processing payment for session {session_id}, order {metadata.get('order_id')}")
                await process_payment_confirmed(session_id, user_id=user_id)
        
        return {"status": "success"}
    except stripe.error.SignatureVerificationError as e:
        logger.error(f"[WEBHOOK] Signature verification failed: {e}")
        raise HTTPException(status_code=400, detail="Invalid signature")
    except Exception as e:
        logger.error(f"[WEBHOOK] Error: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))

@router.post("/payments/process-influencer-payouts")
async def trigger_influencer_payouts(user: User = Depends(get_current_user)):
    """Admin endpoint to trigger processing of due influencer payouts."""
    await require_role(user, ["admin", "super_admin"])
    await process_influencer_scheduled_payouts()
    return {"status": "ok", "message": "Influencer payouts processed"}

@router.get("/payments/scheduled-payouts")
async def get_scheduled_payouts(user: User = Depends(get_current_user)):
    """Admin: view all scheduled payouts."""
    await require_role(user, ["admin", "super_admin"])
    payouts = await db.scheduled_payouts.find({}, {"_id": 0}).sort("due_date", 1).to_list(200)
    return payouts


@router.post("/payments/refund/{order_id}")
async def refund_order(order_id: str, request: Request, user: User = Depends(get_current_user)):
    """
    Admin: refund an order (full or partial).
    Scenarios: 
      1) Full refund within 7 days → reverse all commissions
      2) Partial refund → recalculate proportionally
      3) Refund after influencer paid (D+16) → negative balance (clawback)
    """
    await require_role(user, ["admin", "super_admin"])
    body = await request.json() if request.headers.get("content-type") == "application/json" else {}
    refund_type = body.get("type", "full")  # "full" or "partial"
    partial_amount = body.get("amount", 0)
    
    order = await db.orders.find_one({"order_id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order.get("status") in ("refunded", "cancelled"):
        raise HTTPException(status_code=400, detail="Order already refunded/cancelled")
    
    total_amount = order.get("total_amount", 0)
    refund_amount = total_amount if refund_type == "full" else min(partial_amount, total_amount)
    refund_ratio = refund_amount / total_amount if total_amount > 0 else 1
    now_iso = datetime.now(timezone.utc).isoformat()
    
    # 1. Stripe refund
    session_id = order.get("payment_session_id")
    stripe_refund_id = None
    if session_id:
        try:
            session = stripe.checkout.Session.retrieve(session_id)
            pi_id = session.payment_intent
            if pi_id:
                params = {"payment_intent": pi_id, "metadata": {"order_id": order_id, "type": refund_type}}
                if refund_type == "partial":
                    params["amount"] = int(refund_amount * 100)
                ref = stripe.Refund.create(**params)
                stripe_refund_id = ref.id
                logger.info(f"[REFUND] Stripe refund {ref.id}: {refund_amount} ({refund_type})")
        except stripe.error.StripeError as e:
            logger.error(f"[REFUND] Stripe error: {e}")
    
    # 2. Handle influencer payouts
    cancelled_payouts = 0
    clawback_amount = 0
    
    if refund_type == "full":
        # Cancel ALL pending influencer payouts
        result = await db.scheduled_payouts.update_many(
            {"order_id": order_id, "status": "scheduled"},
            {"$set": {"status": "cancelled", "cancel_reason": "order_refunded", "updated_at": now_iso}}
        )
        cancelled_payouts = result.modified_count
        
        # Check if any were already paid → create clawback
        paid_payouts = await db.scheduled_payouts.find({"order_id": order_id, "status": "paid"}, {"_id": 0}).to_list(10)
        for pp in paid_payouts:
            clawback_amount += pp.get("amount", 0)
            await db.scheduled_payouts.update_one(
                {"payout_id": pp["payout_id"]},
                {"$set": {"status": "clawback", "clawback_amount": pp["amount"], "updated_at": now_iso}}
            )
            # Add negative balance to influencer
            await db.influencers.update_one(
                {"influencer_id": pp["influencer_id"]},
                {"$inc": {"pending_payout_usd": -pp["amount"]}}
            )
            logger.info(f"[REFUND] Clawback {pp['amount']} from influencer {pp['influencer_id']}")
    
    elif refund_type == "partial":
        # Recalculate influencer payouts proportionally
        scheduled = await db.scheduled_payouts.find({"order_id": order_id, "status": "scheduled"}, {"_id": 0}).to_list(10)
        for sp in scheduled:
            new_amount = round(sp["amount"] * (1 - refund_ratio), 2)
            if new_amount <= 0:
                await db.scheduled_payouts.update_one({"payout_id": sp["payout_id"]}, {"$set": {"status": "cancelled", "cancel_reason": "partial_refund", "updated_at": now_iso}})
                cancelled_payouts += 1
            else:
                await db.scheduled_payouts.update_one({"payout_id": sp["payout_id"]}, {"$set": {"amount": new_amount, "updated_at": now_iso}})
    
    # 3. Update order status
    new_status = "refunded" if refund_type == "full" else "partially_refunded"
    await db.orders.update_one(
        {"order_id": order_id},
        {"$set": {"status": new_status, "refund_amount": refund_amount, "refund_type": refund_type, "refunded_at": now_iso, "updated_at": now_iso}}
    )
    await db.payment_transactions.update_one(
        {"order_id": order_id},
        {"$set": {"status": new_status, "updated_at": now_iso}}
    )
    
    # 4. Commission transaction record
    commission_data = order.get("commission_data", {})
    await db.commission_transactions.insert_one({
        "transaction_id": f"ctx_{uuid.uuid4().hex[:12]}",
        "order_id": order_id,
        "event_type": "refund",
        "refund_type": refund_type,
        "refund_amount": refund_amount,
        "refund_ratio": refund_ratio,
        "stripe_refund_id": stripe_refund_id,
        "cancelled_payouts": cancelled_payouts,
        "clawback_amount": clawback_amount,
        "original_commission_data": commission_data,
        "currency": order.get("currency", "EUR"),
        "created_at": now_iso,
    })
    
    # 5. Ledger event
    await write_ledger_event(
        db, event_type="refund", order_id=order_id,
        currency=order.get("currency", "EUR"),
        product_subtotal=-refund_amount,
        buyer_id=order.get("user_id", ""),
        buyer_country=order.get("country", ""),
        status="completed",
        extra={"refund_type": refund_type, "cancelled_payouts": cancelled_payouts, "clawback": clawback_amount},
    )
    
    return {"status": new_status, "refund_amount": refund_amount, "cancelled_payouts": cancelled_payouts, "clawback_amount": clawback_amount}


@router.get("/admin/commission-audit")
async def get_commission_audit(
    user: User = Depends(get_current_user),
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = 500,
):
    """Admin: get all commission transactions for audit trail."""
    await require_role(user, ["admin", "super_admin"])
    query = {}
    if date_from:
        query["created_at"] = {"$gte": date_from}
    if date_to:
        query.setdefault("created_at", {})["$lte"] = date_to
    
    txns = await db.commission_transactions.find(query, {"_id": 0}).sort("created_at", -1).to_list(limit)
    return {"transactions": txns, "count": len(txns)}


@router.get("/admin/export/commission-audit")
async def export_commission_audit(
    user: User = Depends(get_current_user),
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
):
    """Admin: export commission audit trail as CSV."""
    await require_role(user, ["admin", "super_admin"])
    
    query = {}
    if date_from:
        query["created_at"] = {"$gte": date_from}
    if date_to:
        query.setdefault("created_at", {})["$lte"] = date_to
    
    # Get all commission-related data: orders with commission_data
    orders = await db.orders.find(
        {**query, "commission_data": {"$exists": True}},
        {"_id": 0}
    ).sort("created_at", -1).to_list(5000)
    
    import csv
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Date", "Order ID", "Status", "Currency", "Total Amount",
        "Seller ID", "Seller Plan", "Platform Rate", "Platform Gross",
        "Seller Payout", "Influencer ID", "Influencer Tier", "Influencer Rate",
        "Influencer Cut", "Platform Net", "USD Equivalent", "Refund Amount"
    ])
    
    from services.ledger import EXCHANGE_RATES_TO_USD
    
    for order in orders:
        cd = order.get("commission_data", {})
        currency = order.get("currency", "EUR")
        usd_rate = EXCHANGE_RATES_TO_USD.get(currency.upper(), 1.0)
        refund_amt = order.get("refund_amount", 0)
        
        for split in cd.get("splits", []):
            writer.writerow([
                (order.get("created_at", ""))[:10],
                order.get("order_id", ""),
                order.get("status", ""),
                currency,
                order.get("total_amount", 0),
                split.get("seller_id", ""),
                split.get("seller_plan", ""),
                split.get("platform_rate_snapshot", ""),
                split.get("platform_gross", ""),
                split.get("seller_payout", ""),
                split.get("influencer_id", ""),
                split.get("influencer_tier_snapshot", ""),
                split.get("influencer_rate_snapshot", ""),
                split.get("influencer_cut", ""),
                split.get("platform_net", ""),
                round(order.get("total_amount", 0) * usd_rate, 2),
                refund_amt,
            ])
    
    output.seek(0)
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=commission_audit_{datetime.now(timezone.utc).strftime('%Y%m%d')}.csv"}
    )


@router.get("/admin/financial-ledger")
async def get_financial_ledger(
    user: User = Depends(get_current_user),
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    event_type: Optional[str] = None,
    seller_country: Optional[str] = None,
    buyer_country: Optional[str] = None,
    limit: int = 200,
):
    """Admin: view financial ledger entries with filters."""
    await require_role(user, ["admin", "super_admin"])
    
    query = {}
    if date_from:
        query["created_at"] = {"$gte": date_from}
    if date_to:
        query.setdefault("created_at", {})["$lte"] = date_to
    if event_type:
        query["event_type"] = event_type
    if seller_country:
        query["seller_country"] = seller_country.upper()
    if buyer_country:
        query["buyer_country"] = buyer_country.upper()
    
    entries = await db.financial_ledger.find(query, {"_id": 0}).sort("created_at", -1).to_list(limit)
    
    # Summary
    total_gross = sum(e.get("product_subtotal", 0) for e in entries if e.get("event_type") == "order_paid")
    total_platform_fee = sum(e.get("platform_fee", 0) for e in entries if e.get("event_type") == "seller_transfer")
    total_seller_net = sum(e.get("seller_net", 0) for e in entries if e.get("event_type") == "seller_transfer")
    total_usd = sum(e.get("usd_equivalent", 0) for e in entries if e.get("event_type") == "order_paid")
    
    return {
        "entries": entries,
        "summary": {
            "total_gross": round(total_gross, 2),
            "total_platform_fee": round(total_platform_fee, 2),
            "total_seller_net": round(total_seller_net, 2),
            "total_usd_equivalent": round(total_usd, 2),
            "entry_count": len(entries)
        }
    }


@router.get("/admin/export/financial-report")
async def export_financial_report(
    user: User = Depends(get_current_user),
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
):
    """Admin: export financial report as Excel with regional sheets."""
    await require_role(user, ["admin", "super_admin"])
    
    from openpyxl import Workbook
    from fastapi.responses import StreamingResponse
    
    query = {}
    if date_from:
        query["created_at"] = {"$gte": date_from}
    if date_to:
        query.setdefault("created_at", {})["$lte"] = date_to
    
    entries = await db.financial_ledger.find(query, {"_id": 0}).sort("created_at", 1).to_list(5000)

    # Build lookup maps for seller/influencer names
    seller_ids = {e.get("seller_id") for e in entries if e.get("seller_id")}
    influencer_ids = {e.get("influencer_id") for e in entries if e.get("influencer_id")}

    seller_map = {}
    for doc in await db.users.find({"user_id": {"$in": list(seller_ids)}}, {"_id": 0, "user_id": 1, "full_name": 1, "email": 1}).to_list(500):
        seller_map[doc["user_id"]] = f"{doc.get('full_name', '')} <{doc.get('email', '')}>"

    influencer_map = {}
    for doc in await db.influencers.find({"influencer_id": {"$in": list(influencer_ids)}}, {"_id": 0, "influencer_id": 1, "full_name": 1, "email": 1}).to_list(500):
        influencer_map[doc["influencer_id"]] = f"{doc.get('full_name', '')} <{doc.get('email', '')}>"

    wb = Workbook()
    headers = [
        "Fecha", "Order ID", "Tipo de Evento", "Vendedor (Nombre)", "Vendedor (ID)", "País Vendedor",
        "País Comprador", "Región Comprador", "Bruto (€)", "Impuesto Producto", "Tipo Impuesto",
        "Comisión Plataforma", "Impuesto Plataforma", "Fee Stripe", "Neto Vendedor (€)",
        "Influencer (Nombre)", "Influencer (ID)", "Comisión Influencer (€)",
        "Moneda", "Equiv. USD", "Tasa IVA Aplicada", "Cargo Inverso", "Transfer ID", "Estado"
    ]

    def write_sheet(ws, rows):
        ws.append(headers)
        for e in rows:
            sid = e.get("seller_id", "")
            iid = e.get("influencer_id", "")
            ws.append([
                e.get("created_at", "")[:10],
                e.get("order_id", ""),
                e.get("event_type", ""),
                seller_map.get(sid, ""),
                sid,
                e.get("seller_country", ""),
                e.get("buyer_country", ""),
                e.get("buyer_state", ""),
                e.get("product_subtotal", 0),
                e.get("product_tax_amount", 0),
                e.get("product_tax_type", ""),
                e.get("platform_fee", 0),
                e.get("platform_tax_amount", 0),
                e.get("stripe_fee", 0),
                e.get("seller_net", 0),
                influencer_map.get(iid, ""),
                iid,
                e.get("influencer_amount", 0),
                e.get("currency", ""),
                e.get("usd_equivalent", 0),
                e.get("vat_rate_applied", 0),
                "Sí" if e.get("reverse_charge_applied") else "No",
                e.get("transfer_id", ""),
                e.get("status", ""),
            ])

    # Sheet 1 — Full Ledger
    ws1 = wb.active
    ws1.title = "Ledger Completo"
    write_sheet(ws1, entries)

    # Sheet 2 — EU
    from services.ledger import EU_VAT_RATES
    eu_countries = set(EU_VAT_RATES.keys())
    ws2 = wb.create_sheet("Resumen EU")
    write_sheet(ws2, [e for e in entries if e.get("buyer_country") in eu_countries])

    # Sheet 3 — US
    ws3 = wb.create_sheet("Resumen US")
    write_sheet(ws3, [e for e in entries if e.get("buyer_country") == "US"])

    # Sheet 4 — KR
    ws4 = wb.create_sheet("Resumen KR")
    write_sheet(ws4, [e for e in entries if e.get("buyer_country") == "KR"])

    # Sheet 5 — Influencer Commissions Detail
    ws5 = wb.create_sheet("Comisiones Influencers")
    ws5.append(["Fecha", "Influencer (Nombre)", "Influencer (ID)", "Order ID", "Comisión (€)", "Moneda", "Tier", "Estado"])
    inf_entries = [e for e in entries if e.get("influencer_id") and e.get("influencer_amount", 0) > 0]
    for e in inf_entries:
        iid = e.get("influencer_id", "")
        ws5.append([
            e.get("created_at", "")[:10],
            influencer_map.get(iid, ""),
            iid,
            e.get("order_id", ""),
            e.get("influencer_amount", 0),
            e.get("currency", ""),
            e.get("influencer_tier", ""),
            e.get("status", ""),
        ])

    # Sheet 6 — Consolidado USD
    ws6 = wb.create_sheet("Consolidado USD")
    ws6.append(["Fecha", "Order ID", "Tipo Evento", "Equiv. USD", "Moneda Original", "Importe Original", "Tasa de Cambio"])
    for e in entries:
        ws6.append([
            e.get("created_at", "")[:10],
            e.get("order_id", ""),
            e.get("event_type", ""),
            e.get("usd_equivalent", 0),
            e.get("currency", ""),
            e.get("product_subtotal", 0),
            e.get("exchange_rate_to_usd", 0),
        ])

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    period = ""
    if date_from:
        period += f"_{date_from}"
    if date_to:
        period += f"_al_{date_to}"
    if not period:
        period = f"_{datetime.now(timezone.utc).strftime('%Y%m%d')}"

    filename = f"hispaloshop_contabilidad{period}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# Orders
@router.get("/orders")
async def get_orders(user: User = Depends(get_current_user)):
    query = {"user_id": user.user_id}
    if user.role == "producer":
        orders = await db.orders.find({"line_items.producer_id": user.user_id}, {"_id": 0}).to_list(100)
        return orders
    elif user.role == "admin":
        orders = await db.orders.find({}, {"_id": 0}).to_list(1000)
        return orders
    else:
        orders = await db.orders.find(query, {"_id": 0}).to_list(100)
        return orders

@router.get("/orders/{order_id}")
async def get_order(order_id: str, user: User = Depends(get_current_user)):
    order = await db.orders.find_one({"order_id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if user.role == "customer" and order["user_id"] != user.user_id:
        raise HTTPException(status_code=403, detail="Not authorized")
    return order


@router.get("/orders/{order_id}/tracking")
async def get_order_tracking(order_id: str, user: User = Depends(get_current_user)):
    """Legacy tracking endpoint for frontend compatibility."""
    order = await get_order(order_id, user)
    return {
        "order_id": order.get("order_id"),
        "status": order.get("status"),
        "tracking_number": order.get("tracking_number"),
        "tracking_url": order.get("tracking_url"),
        "shipping_carrier": order.get("shipping_carrier"),
        "updated_at": order.get("updated_at"),
    }


@router.post("/orders/{order_id}/cancel")
async def cancel_order(order_id: str, user: User = Depends(get_current_user)):
    """Legacy cancel endpoint for frontend compatibility."""
    order = await get_order(order_id, user)
    current_status = (order.get("status") or "").lower()
    if current_status in {"delivered", "cancelled", "refunded"}:
        raise HTTPException(status_code=400, detail="Order cannot be cancelled")

    await db.orders.update_one(
        {"order_id": order_id},
        {
            "$set": {
                "status": "cancelled",
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }
        },
    )
    return {"success": True, "status": "cancelled"}


@router.post("/orders/{order_id}/reorder")
async def reorder_order(order_id: str, user: User = Depends(get_current_user)):
    """Legacy reorder endpoint for frontend compatibility."""
    order = await get_order(order_id, user)
    if user.role == "customer" and order.get("user_id") != user.user_id:
        raise HTTPException(status_code=403, detail="Not authorized")

    line_items = order.get("line_items", [])
    if not line_items:
        return {"success": False, "reason": "empty_order"}

    cart = await db.carts.find_one({"user_id": user.user_id, "status": "active"})
    cart_items = list(cart.get("items", [])) if cart else []

    for item in line_items:
        product_id = item.get("product_id")
        quantity = int(item.get("quantity", 1) or 1)
        if not product_id or quantity <= 0:
            continue

        unit_price = int(item.get("price_cents", 0) or item.get("unit_price_cents", 0) or 0)
        cart_items.append({
            "product_id": product_id,
            "product_name": item.get("name") or item.get("product_name"),
            "product_image": item.get("image") or item.get("product_image"),
            "seller_id": item.get("producer_id") or item.get("seller_id"),
            "seller_type": "producer",
            "quantity": quantity,
            "unit_price_cents": unit_price,
            "total_price_cents": unit_price * quantity,
            "variant_id": item.get("variant_id"),
            "added_at": datetime.now(timezone.utc),
        })

    if cart:
        await db.carts.update_one(
            {"_id": cart["_id"]},
            {"$set": {"items": cart_items, "updated_at": datetime.now(timezone.utc)}},
        )
    else:
        await db.carts.insert_one({
            "user_id": user.user_id,
            "tenant_id": getattr(user, 'country', None) or "ES",
            "status": "active",
            "items": cart_items,
            "created_at": datetime.now(timezone.utc),
            "updated_at": datetime.now(timezone.utc),
            "expires_at": datetime.now(timezone.utc) + timedelta(days=7),
        })

    return {"success": True, "items_added": len(line_items)}

@router.put("/orders/{order_id}/status")
async def update_order_status(order_id: str, update: OrderStatusUpdate, user: User = Depends(get_current_user)):
    """Update order status with tracking info and send email notification"""
    await require_role(user, ["admin", "producer", "importer"])
    
    # Get the order
    order = await db.orders.find_one({"order_id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    current_status = order.get("status", "pending")
    
    # Validation: shipped requires paid status
    if update.status == "shipped" and current_status not in ["paid", "confirmed", "preparing"]:
        raise HTTPException(status_code=400, detail=f"Cannot mark as shipped. Current status is '{current_status}'. Order must be paid first.")
    
    # Verify producer owns products in this order (if producer)
    producer_ids_in_order = set()
    if user.role == "producer":
        producer_products = await db.products.find({"producer_id": user.user_id}, {"product_id": 1}).to_list(1000)
        producer_product_ids = {p["product_id"] for p in producer_products}
        order_product_ids = {item.get("product_id") for item in order.get("line_items", [])}
        if not order_product_ids.intersection(producer_product_ids):
            raise HTTPException(status_code=403, detail="Not authorized to update this order")
    
    # Get all producer IDs in this order for seller notification
    for item in order.get("line_items", []):
        if item.get("producer_id"):
            producer_ids_in_order.add(item["producer_id"])
    
    # Build status history entry
    status_entry = {
        "status": update.status,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "updated_by": user.user_id,
        "notes": update.notes
    }
    
    # Update order
    update_data = {
        "status": update.status,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    if update.tracking_number:
        update_data["tracking_number"] = update.tracking_number
    if update.tracking_url:
        update_data["tracking_url"] = update.tracking_url
    if update.shipping_carrier:
        update_data["shipping_carrier"] = update.shipping_carrier
    
    # Add shipped_at timestamp when marking as shipped
    if update.status == "shipped":
        update_data["shipped_at"] = datetime.now(timezone.utc).isoformat()
    elif update.status == "delivered":
        update_data["delivered_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.orders.update_one(
        {"order_id": order_id},
        {
            "$set": update_data,
            "$push": {"status_history": status_entry}
        }
    )
    
    # Refresh order with updates
    order.update(update_data)
    
    # Send email notification to customer
    try:
        await send_order_status_email(order, update.status, update.tracking_number, update.tracking_url, update.shipping_carrier)
    except Exception as e:
        logger.error(f"[ORDER_STATUS] Failed to send customer email: {e}")
    
    # Send confirmation email to seller(s) when shipped
    if update.status == "shipped":
        for producer_id in producer_ids_in_order:
            try:
                await send_seller_shipped_email(order, producer_id, update.tracking_number, update.shipping_carrier)
            except Exception as e:
                logger.error(f"[ORDER_STATUS] Failed to send seller email: {e}")
    
    return {"message": "Order status updated", "status": update.status}


async def send_order_status_email(order: dict, new_status: str, tracking_number: str = None, tracking_url: str = None, shipping_carrier: str = None):
    """Send email notification when order status changes - MULTILINGUAL"""
    
    # Get user's preferred language
    user = await db.users.find_one({"user_id": order.get("user_id")}, {"_id": 0, "language": 1})
    user_lang = user.get("language", "es") if user else "es"
    
    # Multilingual status messages (11 languages)
    status_messages = {
        "es": {
            "confirmed": {"title": "Pedido Confirmado", "message": "Tu pedido ha sido confirmado y está siendo procesado."},
            "preparing": {"title": "Pedido en Preparación", "message": "Tu pedido está siendo preparado para el envío."},
            "shipped": {"title": "Pedido Enviado", "message": "Tu pedido ha sido enviado."},
            "delivered": {"title": "Pedido Entregado", "message": "Tu pedido ha sido entregado. ¡Gracias por comprar en Hispaloshop!"},
            "cancelled": {"title": "Pedido Cancelado", "message": "Tu pedido ha sido cancelado. Si tienes preguntas, contáctanos."},
            "default": {"title": "Actualización de Pedido", "message": f"El estado de tu pedido ha cambiado a: {new_status}"},
            "tracking": "Número de seguimiento",
            "track": "Seguir envío",
            "order": "Pedido",
            "total": "Total",
            "products": "Productos",
            "view_order": "Ver Pedido",
            "rights": "Todos los derechos reservados."
        },
        "en": {
            "confirmed": {"title": "Order Confirmed", "message": "Your order has been confirmed and is being processed."},
            "preparing": {"title": "Order Being Prepared", "message": "Your order is being prepared for shipping."},
            "shipped": {"title": "Order Shipped", "message": "Your order has been shipped."},
            "delivered": {"title": "Order Delivered", "message": "Your order has been delivered. Thank you for shopping at Hispaloshop!"},
            "cancelled": {"title": "Order Cancelled", "message": "Your order has been cancelled. If you have questions, please contact us."},
            "default": {"title": "Order Update", "message": f"Your order status has changed to: {new_status}"},
            "tracking": "Tracking Number",
            "track": "Track shipment",
            "order": "Order",
            "total": "Total",
            "products": "Products",
            "view_order": "View Order",
            "rights": "All rights reserved."
        },
        "fr": {
            "confirmed": {"title": "Commande Confirmée", "message": "Votre commande a été confirmée et est en cours de traitement."},
            "preparing": {"title": "Commande en Préparation", "message": "Votre commande est en cours de préparation pour l'expédition."},
            "shipped": {"title": "Commande Expédiée", "message": "Votre commande a été expédiée."},
            "delivered": {"title": "Commande Livrée", "message": "Votre commande a été livrée. Merci d'avoir fait vos achats chez Hispaloshop!"},
            "cancelled": {"title": "Commande Annulée", "message": "Votre commande a été annulée. Si vous avez des questions, contactez-nous."},
            "default": {"title": "Mise à jour de la Commande", "message": f"Le statut de votre commande a changé en: {new_status}"},
            "tracking": "Numéro de suivi",
            "track": "Suivre l'envoi",
            "order": "Commande",
            "total": "Total",
            "products": "Produits",
            "view_order": "Voir la Commande",
            "rights": "Tous droits réservés."
        },
        "de": {
            "confirmed": {"title": "Bestellung Bestätigt", "message": "Ihre Bestellung wurde bestätigt und wird bearbeitet."},
            "preparing": {"title": "Bestellung wird Vorbereitet", "message": "Ihre Bestellung wird für den Versand vorbereitet."},
            "shipped": {"title": "Bestellung Versendet", "message": "Ihre Bestellung wurde versendet."},
            "delivered": {"title": "Bestellung Geliefert", "message": "Ihre Bestellung wurde geliefert. Vielen Dank für Ihren Einkauf bei Hispaloshop!"},
            "cancelled": {"title": "Bestellung Storniert", "message": "Ihre Bestellung wurde storniert. Bei Fragen kontaktieren Sie uns."},
            "default": {"title": "Bestellungsupdate", "message": f"Der Status Ihrer Bestellung hat sich geändert zu: {new_status}"},
            "tracking": "Sendungsnummer",
            "track": "Sendung verfolgen",
            "order": "Bestellung",
            "total": "Gesamt",
            "products": "Produkte",
            "view_order": "Bestellung Anzeigen",
            "rights": "Alle Rechte vorbehalten."
        },
        "pt": {
            "confirmed": {"title": "Pedido Confirmado", "message": "Seu pedido foi confirmado e está sendo processado."},
            "preparing": {"title": "Pedido em Preparação", "message": "Seu pedido está sendo preparado para envio."},
            "shipped": {"title": "Pedido Enviado", "message": "Seu pedido foi enviado."},
            "delivered": {"title": "Pedido Entregue", "message": "Seu pedido foi entregue. Obrigado por comprar na Hispaloshop!"},
            "cancelled": {"title": "Pedido Cancelado", "message": "Seu pedido foi cancelado. Se tiver dúvidas, entre em contato."},
            "default": {"title": "Atualização do Pedido", "message": f"O status do seu pedido mudou para: {new_status}"},
            "tracking": "Número de rastreamento",
            "track": "Rastrear envio",
            "order": "Pedido",
            "total": "Total",
            "products": "Produtos",
            "view_order": "Ver Pedido",
            "rights": "Todos os direitos reservados."
        },
        "ar": {
            "confirmed": {"title": "تم تأكيد الطلب", "message": "تم تأكيد طلبك وجاري معالجته."},
            "preparing": {"title": "الطلب قيد التحضير", "message": "يتم تحضير طلبك للشحن."},
            "shipped": {"title": "تم شحن الطلب", "message": "تم شحن طلبك."},
            "delivered": {"title": "تم تسليم الطلب", "message": "تم تسليم طلبك. شكراً للتسوق في Hispaloshop!"},
            "cancelled": {"title": "تم إلغاء الطلب", "message": "تم إلغاء طلبك. إذا كان لديك أسئلة، تواصل معنا."},
            "default": {"title": "تحديث الطلب", "message": f"تغيرت حالة طلبك إلى: {new_status}"},
            "tracking": "رقم التتبع",
            "track": "تتبع الشحنة",
            "order": "الطلب",
            "total": "المجموع",
            "products": "المنتجات",
            "view_order": "عرض الطلب",
            "rights": "جميع الحقوق محفوظة."
        },
        "hi": {
            "confirmed": {"title": "ऑर्डर की पुष्टि हो गई", "message": "आपके ऑर्डर की पुष्टि हो गई है और प्रोसेस किया जा रहा है।"},
            "preparing": {"title": "ऑर्डर तैयार हो रहा है", "message": "आपका ऑर्डर शिपिंग के लिए तैयार किया जा रहा है।"},
            "shipped": {"title": "ऑर्डर भेज दिया गया", "message": "आपका ऑर्डर भेज दिया गया है।"},
            "delivered": {"title": "ऑर्डर डिलीवर हो गया", "message": "आपका ऑर्डर डिलीवर हो गया है। Hispaloshop पर खरीदारी के लिए धन्यवाद!"},
            "cancelled": {"title": "ऑर्डर रद्द", "message": "आपका ऑर्डर रद्द कर दिया गया है। प्रश्न हों तो संपर्क करें।"},
            "default": {"title": "ऑर्डर अपडेट", "message": f"आपके ऑर्डर की स्थिति बदल गई है: {new_status}"},
            "tracking": "ट्रैकिंग नंबर",
            "track": "शिपमेंट ट्रैक करें",
            "order": "ऑर्डर",
            "total": "कुल",
            "products": "उत्पाद",
            "view_order": "ऑर्डर देखें",
            "rights": "सर्वाधिकार सुरक्षित।"
        },
        "ja": {
            "confirmed": {"title": "注文確認済み", "message": "ご注文が確認され、処理中です。"},
            "preparing": {"title": "注文準備中", "message": "ご注文は発送準備中です。"},
            "shipped": {"title": "注文発送済み", "message": "ご注文は発送されました。"},
            "delivered": {"title": "注文配達済み", "message": "ご注文が配達されました。Hispaloshopでのお買い物ありがとうございます！"},
            "cancelled": {"title": "注文キャンセル", "message": "ご注文はキャンセルされました。ご質問があればお問い合わせください。"},
            "default": {"title": "注文更新", "message": f"ご注文のステータスが変更されました: {new_status}"},
            "tracking": "追跡番号",
            "track": "配送を追跡",
            "order": "注文",
            "total": "合計",
            "products": "商品",
            "view_order": "注文を見る",
            "rights": "全著作権所有。"
        },
        "ko": {
            "confirmed": {"title": "주문 확인됨", "message": "주문이 확인되었으며 처리 중입니다."},
            "preparing": {"title": "주문 준비 중", "message": "주문이 배송 준비 중입니다."},
            "shipped": {"title": "주문 배송됨", "message": "주문이 배송되었습니다."},
            "delivered": {"title": "주문 배달됨", "message": "주문이 배달되었습니다. Hispaloshop에서 쇼핑해 주셔서 감사합니다!"},
            "cancelled": {"title": "주문 취소됨", "message": "주문이 취소되었습니다. 문의사항이 있으시면 연락해 주세요."},
            "default": {"title": "주문 업데이트", "message": f"주문 상태가 변경되었습니다: {new_status}"},
            "tracking": "추적 번호",
            "track": "배송 추적",
            "order": "주문",
            "total": "총액",
            "products": "상품",
            "view_order": "주문 보기",
            "rights": "모든 권리 보유."
        },
        "ru": {
            "confirmed": {"title": "Заказ Подтвержден", "message": "Ваш заказ подтвержден и обрабатывается."},
            "preparing": {"title": "Заказ Готовится", "message": "Ваш заказ готовится к отправке."},
            "shipped": {"title": "Заказ Отправлен", "message": "Ваш заказ отправлен."},
            "delivered": {"title": "Заказ Доставлен", "message": "Ваш заказ доставлен. Спасибо за покупку в Hispaloshop!"},
            "cancelled": {"title": "Заказ Отменен", "message": "Ваш заказ отменен. Если есть вопросы, свяжитесь с нами."},
            "default": {"title": "Обновление Заказа", "message": f"Статус вашего заказа изменен на: {new_status}"},
            "tracking": "Номер отслеживания",
            "track": "Отследить отправление",
            "order": "Заказ",
            "total": "Итого",
            "products": "Товары",
            "view_order": "Посмотреть Заказ",
            "rights": "Все права защищены."
        },
        "zh": {
            "confirmed": {"title": "订单已确认", "message": "您的订单已确认，正在处理中。"},
            "preparing": {"title": "订单准备中", "message": "您的订单正在准备发货。"},
            "shipped": {"title": "订单已发货", "message": "您的订单已发货。"},
            "delivered": {"title": "订单已送达", "message": "您的订单已送达。感谢您在Hispaloshop购物！"},
            "cancelled": {"title": "订单已取消", "message": "您的订单已取消。如有疑问，请联系我们。"},
            "default": {"title": "订单更新", "message": f"您的订单状态已更改为：{new_status}"},
            "tracking": "追踪号码",
            "track": "追踪货物",
            "order": "订单",
            "total": "总计",
            "products": "商品",
            "view_order": "查看订单",
            "rights": "版权所有。"
        }
    }
    
    # Fallback to Spanish if language not supported
    lang_messages = status_messages.get(user_lang, status_messages["es"])
    status_info = lang_messages.get(new_status, lang_messages["default"])
    
    # Build tracking section if available
    tracking_html = ""
    if tracking_number or shipping_carrier:
        carrier_text = f"<p style='margin: 0 0 5px 0; color: #7A7A7A;'>Transportista: <strong>{shipping_carrier}</strong></p>" if shipping_carrier else ""
        tracking_html = f"""
        <div style="background: #e8f5e9; padding: 20px; border-radius: 8px; margin: 20px 0; border-left: 4px solid #4CAF50;">
            <p style="margin: 0 0 5px 0; font-weight: bold; color: #2e7d32;">📦 Información de Envío</p>
            {carrier_text}
            {f"<p style='margin: 5px 0;'><strong>{lang_messages['tracking']}:</strong></p><p style='margin: 0; font-size: 18px; font-family: monospace; background: white; padding: 10px; border-radius: 4px;'>{tracking_number}</p>" if tracking_number else ""}
            {"<a href='" + tracking_url + "' style='display: inline-block; margin-top: 10px; padding: 8px 16px; background: #4CAF50; color: white; text-decoration: none; border-radius: 4px;'>" + lang_messages['track'] + " →</a>" if tracking_url else ""}
        </div>
        """
    
    # Build order items summary
    items_html = ""
    for item in order.get("line_items", [])[:5]:  # Show max 5 items
        items_html += f"<li>{item.get('product_name', 'Product')} x {item.get('quantity', 1)}</li>"
    
    email_html = f"""
    <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
        <div style="background: #1C1C1C; color: white; padding: 20px; text-align: center;">
            <h1 style="margin: 0; font-size: 24px;">Hispaloshop</h1>
        </div>
        
        <div style="padding: 30px;">
            <h2 style="color: #1C1C1C; margin-bottom: 10px;">{status_info['title']}</h2>
            <p style="color: #4A4A4A;">{status_info['message']}</p>
            
            <div style="background: #FAF7F2; padding: 20px; border-radius: 8px; margin: 20px 0;">
                <p style="margin: 0 0 10px 0;"><strong>{lang_messages['order']}:</strong> #{order.get('order_id', '')[:8]}</p>
                <p style="margin: 0;"><strong>{lang_messages['total']}:</strong> €{order.get('total_amount', 0):.2f}</p>
            </div>
            
            {tracking_html}
            
            <div style="margin: 20px 0;">
                <p style="font-weight: bold; margin-bottom: 10px;">{lang_messages['products']}:</p>
                <ul style="color: #4A4A4A; padding-left: 20px;">
                    {items_html}
                </ul>
            </div>
            
            <div style="text-align: center; margin-top: 30px;">
                <a href="{FRONTEND_URL}/dashboard/orders" 
                   style="display: inline-block; padding: 12px 30px; background: #1C1C1C; color: white; text-decoration: none; border-radius: 25px;">
                    {lang_messages['view_order']}
                </a>
            </div>
        </div>
        
        <div style="background: #f5f5f5; padding: 20px; text-align: center; font-size: 12px; color: #7A7A7A;">
            <p>© 2026 Hispaloshop. {lang_messages['rights']}</p>
        </div>
    </div>
    """
    
    send_email(
        to=order.get("user_email"),
        subject=f"Hispaloshop - {status_info['title']} - #{order.get('order_id', '')[:8]}",
        html=email_html
    )
    logger.info(f"[ORDER_STATUS] Email sent to {order.get('user_email')} for order {order.get('order_id')} in {user_lang}")


async def send_new_order_email_to_producer(producer_id: str, order: dict, producer_items: list):
    """Send email notification to producer when they receive a new order"""
    from services.subscriptions import get_seller_commission_rate
    
    # Get producer info
    producer = await db.users.find_one(
        {"user_id": producer_id},
        {"_id": 0, "email": 1, "name": 1, "company_name": 1, "preferred_language": 1, "subscription": 1},
    )
    if not producer or not producer.get("email"):
        logger.warning(f"[EMAIL] Producer {producer_id} not found or no email")
        return
    
    producer_name = producer.get("company_name") or producer.get("name", "Vendedor")
    producer_lang = producer.get("preferred_language", "es")
    seller_plan = str((producer.get("subscription") or {}).get("plan", "FREE")).upper()
    commission_rate = get_seller_commission_rate(seller_plan)
    commission_percent = int(round(commission_rate * 100))
    
    # Calculate total for this producer
    producer_total = sum(item.get("price", 0) * item.get("quantity", 1) for item in producer_items)
    commission = producer_total * commission_rate
    net_earnings = producer_total - commission
    
    # Build items list HTML
    items_html = ""
    for item in producer_items:
        item_total = item.get("price", 0) * item.get("quantity", 1)
        items_html += f"""
        <tr>
            <td style="padding: 12px; border-bottom: 1px solid #eee;">
                <strong>{item.get('product_name', 'Producto')}</strong>
                {f"<br><span style='color: #666; font-size: 12px;'>{item.get('variant_name', '')}</span>" if item.get('variant_name') else ""}
            </td>
            <td style="padding: 12px; border-bottom: 1px solid #eee; text-align: center;">{item.get('quantity', 1)}</td>
            <td style="padding: 12px; border-bottom: 1px solid #eee; text-align: right;">€{item.get('price', 0):.2f}</td>
            <td style="padding: 12px; border-bottom: 1px solid #eee; text-align: right;">€{item_total:.2f}</td>
        </tr>
        """
    
    # Shipping address
    shipping = order.get("shipping_address", {})
    address_html = ""
    if shipping:
        address_html = f"""
        <div style="background: #f8f5f0; padding: 15px; border-radius: 8px; margin: 20px 0; border-left: 4px solid #8B7355;">
            <p style="margin: 0 0 10px 0; font-weight: bold; color: #8B7355;">📍 Dirección de Envío</p>
            <p style="margin: 0; color: #555; line-height: 1.6;">
                <strong>{shipping.get('name', order.get('user_name', ''))}</strong><br>
                {shipping.get('street', '')}<br>
                {shipping.get('postal_code', '')} {shipping.get('city', '')}<br>
                {shipping.get('province', '')}, {shipping.get('country', '')}<br>
                {f"Tel: {shipping.get('phone', '')}" if shipping.get('phone') else ""}
            </p>
        </div>
        """
    
    # Language-specific messages
    lang_messages = {
        "es": {
            "subject": "🎉 ¡Nuevo Pedido Recibido!",
            "greeting": f"¡Hola {producer_name}!",
            "intro": "¡Tienes un nuevo pedido! Un cliente ha realizado una compra de tus productos.",
            "order_number": "Nº Pedido",
            "date": "Fecha",
            "customer": "Cliente",
            "product": "Producto",
            "qty": "Cant.",
            "price": "Precio",
            "total": "Total",
            "subtotal": "Subtotal",
            "commission": f"Comisión plataforma ({commission_percent}%)",
            "net_earnings": "Tu ganancia neta",
            "ship_to": "Enviar a",
            "action_needed": "⚡ Acción Requerida",
            "action_text": "Por favor, prepara el pedido y márcalo como enviado en tu panel de vendedor cuando lo hayas despachado.",
            "go_to_orders": "Ver Mis Pedidos",
            "rights": "Todos los derechos reservados"
        },
        "en": {
            "subject": "🎉 New Order Received!",
            "greeting": f"Hello {producer_name}!",
            "intro": "You have a new order! A customer has purchased your products.",
            "order_number": "Order #",
            "date": "Date",
            "customer": "Customer",
            "product": "Product",
            "qty": "Qty",
            "price": "Price",
            "total": "Total",
            "subtotal": "Subtotal",
            "commission": f"Platform commission ({commission_percent}%)",
            "net_earnings": "Your net earnings",
            "ship_to": "Ship to",
            "action_needed": "⚡ Action Required",
            "action_text": "Please prepare the order and mark it as shipped in your seller dashboard when dispatched.",
            "go_to_orders": "View My Orders",
            "rights": "All rights reserved"
        }
    }
    
    msg = lang_messages.get(producer_lang, lang_messages["es"])
    order_date = datetime.fromisoformat(order.get("created_at", "")).strftime("%d/%m/%Y %H:%M") if order.get("created_at") else ""
    
    email_html = f"""
    <div style="font-family: 'Segoe UI', Arial, sans-serif; max-width: 600px; margin: 0 auto; background: #ffffff;">
        <!-- Header -->
        <div style="background: linear-gradient(135deg, #8B7355 0%, #6B5344 100%); color: white; padding: 30px; text-align: center;">
            <h1 style="margin: 0; font-size: 28px; font-weight: 300;">Hispaloshop</h1>
            <p style="margin: 10px 0 0 0; font-size: 14px; opacity: 0.9;">Panel de Vendedor</p>
        </div>
        
        <!-- Main Content -->
        <div style="padding: 30px;">
            <!-- Celebration Banner -->
            <div style="background: linear-gradient(135deg, #4CAF50 0%, #45a049 100%); color: white; padding: 20px; border-radius: 12px; text-align: center; margin-bottom: 25px;">
                <h2 style="margin: 0; font-size: 22px;">{msg['subject']}</h2>
                <p style="margin: 10px 0 0 0; opacity: 0.9;">{msg['intro']}</p>
            </div>
            
            <!-- Order Info -->
            <div style="display: flex; justify-content: space-between; margin-bottom: 20px; flex-wrap: wrap;">
                <div style="margin-bottom: 10px;">
                    <p style="margin: 0; color: #888; font-size: 12px; text-transform: uppercase;">{msg['order_number']}</p>
                    <p style="margin: 5px 0 0 0; font-size: 18px; font-weight: bold; color: #1C1C1C;">#{order.get('order_id', '')[:8]}</p>
                </div>
                <div style="margin-bottom: 10px;">
                    <p style="margin: 0; color: #888; font-size: 12px; text-transform: uppercase;">{msg['date']}</p>
                    <p style="margin: 5px 0 0 0; font-size: 14px; color: #1C1C1C;">{order_date}</p>
                </div>
                <div style="margin-bottom: 10px;">
                    <p style="margin: 0; color: #888; font-size: 12px; text-transform: uppercase;">{msg['customer']}</p>
                    <p style="margin: 5px 0 0 0; font-size: 14px; color: #1C1C1C;">{order.get('user_name', 'Cliente')}</p>
                </div>
            </div>
            
            <!-- Products Table -->
            <table style="width: 100%; border-collapse: collapse; margin: 20px 0;">
                <thead>
                    <tr style="background: #f8f5f0;">
                        <th style="padding: 12px; text-align: left; color: #8B7355; font-size: 12px; text-transform: uppercase;">{msg['product']}</th>
                        <th style="padding: 12px; text-align: center; color: #8B7355; font-size: 12px; text-transform: uppercase;">{msg['qty']}</th>
                        <th style="padding: 12px; text-align: right; color: #8B7355; font-size: 12px; text-transform: uppercase;">{msg['price']}</th>
                        <th style="padding: 12px; text-align: right; color: #8B7355; font-size: 12px; text-transform: uppercase;">{msg['total']}</th>
                    </tr>
                </thead>
                <tbody>
                    {items_html}
                </tbody>
            </table>
            
            <!-- Totals -->
            <div style="background: #f8f5f0; padding: 15px; border-radius: 8px; margin: 20px 0;">
                <div style="display: flex; justify-content: space-between; margin-bottom: 8px;">
                    <span style="color: #666;">{msg['subtotal']}</span>
                    <span style="color: #1C1C1C;">€{producer_total:.2f}</span>
                </div>
                <div style="display: flex; justify-content: space-between; margin-bottom: 8px;">
                    <span style="color: #666;">{msg['commission']}</span>
                    <span style="color: #e53935;">-€{commission:.2f}</span>
                </div>
                <div style="display: flex; justify-content: space-between; padding-top: 10px; border-top: 2px solid #8B7355;">
                    <span style="font-weight: bold; color: #1C1C1C;">{msg['net_earnings']}</span>
                    <span style="font-weight: bold; font-size: 18px; color: #4CAF50;">€{net_earnings:.2f}</span>
                </div>
            </div>
            
            <!-- Shipping Address -->
            {address_html}
            
            <!-- Action Required -->
            <div style="background: #fff3e0; padding: 15px; border-radius: 8px; margin: 20px 0; border-left: 4px solid #ff9800;">
                <p style="margin: 0 0 10px 0; font-weight: bold; color: #e65100;">{msg['action_needed']}</p>
                <p style="margin: 0; color: #555;">{msg['action_text']}</p>
            </div>
            
            <!-- CTA Button -->
            <div style="text-align: center; margin: 30px 0;">
                <a href="https://auth-rework.preview.emergentagent.com/producer/orders" 
                   style="display: inline-block; background: linear-gradient(135deg, #8B7355 0%, #6B5344 100%); color: white; padding: 15px 30px; text-decoration: none; border-radius: 8px; font-weight: bold; font-size: 16px;">
                    {msg['go_to_orders']} →
                </a>
            </div>
        </div>
        
        <!-- Footer -->
        <div style="background: #f5f5f5; padding: 20px; text-align: center; border-top: 1px solid #eee;">
            <p style="margin: 0; color: #888; font-size: 12px;">© 2026 Hispaloshop. {msg['rights']}</p>
        </div>
    </div>
    """
    
    send_email(
        to=producer.get("email"),
        subject=f"Hispaloshop - {msg['subject']} - #{order.get('order_id', '')[:8]}",
        html=email_html
    )
    logger.info(f"[NEW_ORDER_EMAIL] Sent to producer {producer_id} ({producer.get('email')}) for order {order.get('order_id')}")


async def send_seller_shipped_email(order: dict, producer_id: str, tracking_number: str = None, shipping_carrier: str = None):
    """Send confirmation email to seller when order is marked as shipped"""
    
    # Get seller info
    seller = await db.users.find_one({"user_id": producer_id}, {"_id": 0, "email": 1, "name": 1, "company_name": 1})
    if not seller or not seller.get("email"):
        return
    
    seller_name = seller.get("company_name") or seller.get("name", "Vendedor")
    
    # Get products from this seller in the order
    seller_items = []
    seller_total = 0
    for item in order.get("line_items", []):
        if item.get("producer_id") == producer_id:
            seller_items.append(item)
            seller_total += item.get("price", 0) * item.get("quantity", 1)
    
    if not seller_items:
        return
    
    # Build items list HTML
    items_html = ""
    for item in seller_items:
        items_html += f"""
        <tr>
            <td style="padding: 10px; border-bottom: 1px solid #eee;">{item.get('product_name', 'Producto')}</td>
            <td style="padding: 10px; border-bottom: 1px solid #eee; text-align: center;">{item.get('quantity', 1)}</td>
            <td style="padding: 10px; border-bottom: 1px solid #eee; text-align: right;">€{item.get('price', 0):.2f}</td>
        </tr>
        """
    
    # Tracking info
    tracking_info = ""
    if tracking_number or shipping_carrier:
        tracking_info = f"""
        <div style="background: #e3f2fd; padding: 15px; border-radius: 8px; margin: 20px 0;">
            <p style="margin: 0; font-weight: bold; color: #1565c0;">📦 Datos de Envío Registrados</p>
            {f"<p style='margin: 5px 0 0 0;'>Transportista: <strong>{shipping_carrier}</strong></p>" if shipping_carrier else ""}
            {f"<p style='margin: 5px 0 0 0;'>Nº Seguimiento: <strong>{tracking_number}</strong></p>" if tracking_number else ""}
        </div>
        """
    
    # Shipping address
    shipping = order.get("shipping_address", {})
    address_html = ""
    if shipping:
        address_html = f"""
        <div style="background: #f5f5f5; padding: 15px; border-radius: 8px; margin: 20px 0;">
            <p style="margin: 0 0 10px 0; font-weight: bold;">📍 Dirección de Envío</p>
            <p style="margin: 0; color: #555;">
                {shipping.get('name', '')}<br>
                {shipping.get('street', '')}<br>
                {shipping.get('postal_code', '')} {shipping.get('city', '')}<br>
                {shipping.get('country', '')}
            </p>
        </div>
        """
    
    email_html = f"""
    <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
        <div style="background: #8B7355; color: white; padding: 20px; text-align: center;">
            <h1 style="margin: 0; font-size: 24px;">Hispaloshop</h1>
            <p style="margin: 5px 0 0 0; font-size: 14px; opacity: 0.9;">Panel de Vendedor</p>
        </div>
        
        <div style="padding: 30px;">
            <h2 style="color: #1C1C1C; margin-bottom: 10px;">✅ Envío Registrado Correctamente</h2>
            <p style="color: #4A4A4A;">
                Hola {seller_name},<br><br>
                Has marcado como enviado el pedido <strong>#{order.get('order_id', '')[:8]}</strong>. 
                El cliente ha sido notificado por email.
            </p>
            
            {tracking_info}
            
            <div style="margin: 20px 0;">
                <p style="font-weight: bold; margin-bottom: 10px;">Productos enviados:</p>
                <table style="width: 100%; border-collapse: collapse;">
                    <thead>
                        <tr style="background: #f5f5f5;">
                            <th style="padding: 10px; text-align: left;">Producto</th>
                            <th style="padding: 10px; text-align: center;">Cantidad</th>
                            <th style="padding: 10px; text-align: right;">Precio</th>
                        </tr>
                    </thead>
                    <tbody>
                        {items_html}
                    </tbody>
                    <tfoot>
                        <tr style="background: #FAF7F2;">
                            <td colspan="2" style="padding: 10px; font-weight: bold;">Tu parte (80%)</td>
                            <td style="padding: 10px; text-align: right; font-weight: bold; color: #2e7d32;">€{(seller_total * 0.8):.2f}</td>
                        </tr>
                    </tfoot>
                </table>
            </div>
            
            {address_html}
            
            <div style="text-align: center; margin-top: 30px;">
                <a href="{FRONTEND_URL}/producer/orders" 
                   style="display: inline-block; padding: 12px 30px; background: #8B7355; color: white; text-decoration: none; border-radius: 25px;">
                    Ver Mis Pedidos
                </a>
            </div>
        </div>
        
        <div style="background: #f5f5f5; padding: 20px; text-align: center; font-size: 12px; color: #7A7A7A;">
            <p>© 2026 Hispaloshop. Todos los derechos reservados.</p>
        </div>
    </div>
    """
    
    send_email(
        to=seller["email"],
        subject=f"Hispaloshop - Envío Registrado - Pedido #{order.get('order_id', '')[:8]}",
        html=email_html
    )
    logger.info(f"[ORDER_SHIPPED] Seller confirmation email sent to {seller['email']} for order {order.get('order_id')}")




async def check_and_notify_influencer_withdrawal_available(influencer_id: str, db_ref):
    """Check if influencer has reached withdrawal threshold and send notification."""
    try:
        influencer = await db_ref.influencers.find_one({"influencer_id": influencer_id}, {"_id": 0})
        if not influencer:
            return
        if not influencer.get("stripe_account_id") or not influencer.get("stripe_onboarding_complete"):
            return
        
        now = datetime.now(timezone.utc)
        pending_commissions = await db_ref.influencer_commissions.find({
            "influencer_id": influencer_id,
            "commission_status": "pending"
        }, {"_id": 0}).to_list(1000)
        
        available_balance = 0
        for comm in pending_commissions:
            payment_date_str = comm.get("payment_available_date")
            if payment_date_str:
                try:
                    payment_date = datetime.fromisoformat(payment_date_str.replace('Z', '+00:00'))
                    if payment_date <= now:
                        available_balance += comm.get("commission_amount", 0)
                except:
                    pass
        
        available_balance = round(available_balance, 2)
        if available_balance < 50:
            return
        
        last_notification = influencer.get("last_withdrawal_notification_amount", 0)
        threshold_reached = int(available_balance / 50) * 50
        if threshold_reached <= last_notification:
            return
        
        email = influencer.get("email")
        full_name = influencer.get("full_name", "Influencer")
        
        if email:
            try:
                html = f"""
                <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
                    <h1 style="color: #1C1C1C; text-align: center;">HispaloShop</h1>
                    <div style="background: linear-gradient(135deg, #10b981 0%, #059669 100%); color: white; padding: 30px; border-radius: 12px; text-align: center; margin-bottom: 20px;">
                        <h2 style="margin: 0 0 10px;">Ya puedes retirar tus comisiones</h2>
                        <p style="margin: 0; font-size: 42px; font-weight: bold;">&euro;{available_balance:.2f}</p>
                        <p style="margin: 10px 0 0; opacity: 0.9;">disponibles para retirar</p>
                    </div>
                    <p>Hola {full_name}, has alcanzado <strong>&euro;{available_balance:.2f}</strong> en comisiones disponibles.</p>
                    <div style="text-align: center; margin: 30px 0;">
                        <a href="https://www.hispaloshop.com/influencer/dashboard" 
                           style="display: inline-block; background: #1C1C1C; color: white; padding: 16px 32px; text-decoration: none; border-radius: 8px; font-weight: bold;">
                            Ir a mi panel de influencer
                        </a>
                    </div>
                </div>
                """
                send_email(email, f"Ya puedes retirar &euro;{available_balance:.2f} en comisiones!", html)
                await db_ref.influencers.update_one(
                    {"influencer_id": influencer_id},
                    {"$set": {"last_withdrawal_notification_amount": threshold_reached}}
                )
                logger.info(f"[NOTIFICATION] Sent withdrawal notification to {email} for EUR {available_balance:.2f}")
            except Exception as e:
                logger.error(f"[NOTIFICATION] Failed to send withdrawal notification to {email}: {e}")
    except Exception as e:
        logger.error(f"[NOTIFICATION] Error checking withdrawal for {influencer_id}: {e}")

